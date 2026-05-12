import pandas as pd
import time
import os
from enriquecer import enrich_lead


def _generar_excel_profesional(df, output_file, col_nombre):
    """Genera un Excel con formato profesional y una hoja de resumen."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # Reordenar columnas para que las más importantes queden al inicio
    columnas_inicio = [col_nombre, 'email', 'gerente']
    columnas_restantes = [c for c in df.columns if c not in columnas_inicio]
    columnas_ordenadas = [c for c in columnas_inicio if c in df.columns] + columnas_restantes
    df = df[columnas_ordenadas]

    total = len(df)
    con_email = int(df['email'].notna().sum()) if 'email' in df.columns else 0
    con_gerente = int(df['gerente'].notna().sum()) if 'gerente' in df.columns else 0
    completos = int((df['email'].notna() & df['gerente'].notna()).sum()) if {'email', 'gerente'}.issubset(df.columns) else 0

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Leads Enriquecidos', index=False)

        resumen_df = pd.DataFrame([
            {'Metrica': 'Total de registros', 'Valor': total},
            {'Metrica': 'Con email', 'Valor': con_email},
            {'Metrica': 'Con gerente', 'Valor': con_gerente},
            {'Metrica': 'Completos (email + gerente)', 'Valor': completos},
            {'Metrica': '% con email', 'Valor': (con_email / total if total else 0)},
            {'Metrica': '% con gerente', 'Valor': (con_gerente / total if total else 0)},
            {'Metrica': '% completos', 'Valor': (completos / total if total else 0)},
        ])
        resumen_df.to_excel(writer, sheet_name='Resumen', index=False)

        wb = writer.book
        ws = wb['Leads Enriquecidos']
        ws_resumen = wb['Resumen']

        # Estilos base
        color_header = '1F4E78'
        color_header_resumen = '2F75B5'
        color_ok = 'E2F0D9'
        color_warn = 'FCE4D6'

        header_font = Font(color='FFFFFF', bold=True)
        header_fill = PatternFill(fill_type='solid', fgColor=color_header)
        header_fill_resumen = PatternFill(fill_type='solid', fgColor=color_header_resumen)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )

        # Formato de hoja principal
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = thin_border

        # Colorear estado de datos en email/gerente para lectura rápida
        idx_email = None
        idx_gerente = None
        for idx, name in enumerate(df.columns, start=1):
            if str(name).strip().lower() == 'email':
                idx_email = idx
            if str(name).strip().lower() == 'gerente':
                idx_gerente = idx

        if idx_email:
            for fila in range(2, ws.max_row + 1):
                celda = ws.cell(row=fila, column=idx_email)
                tiene_dato = bool(celda.value and str(celda.value).strip())
                celda.fill = PatternFill(fill_type='solid', fgColor=(color_ok if tiene_dato else color_warn))

        if idx_gerente:
            for fila in range(2, ws.max_row + 1):
                celda = ws.cell(row=fila, column=idx_gerente)
                tiene_dato = bool(celda.value and str(celda.value).strip())
                celda.fill = PatternFill(fill_type='solid', fgColor=(color_ok if tiene_dato else color_warn))

        # Ancho automático de columnas (con límite para que no se desborde)
        for i, col in enumerate(df.columns, start=1):
            max_len = len(str(col))
            for val in df[col].astype(str).fillna(''):
                max_len = max(max_len, len(val))
            adjusted = min(max(12, max_len + 2), 45)
            ws.column_dimensions[get_column_letter(i)].width = adjusted

        # Convertir en tabla para estilo profesional y filtros claros
        if ws.max_row >= 2 and ws.max_column >= 1:
            table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            tabla = Table(displayName='TablaLeads', ref=table_ref)
            tabla.tableStyleInfo = TableStyleInfo(
                name='TableStyleMedium2',
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tabla)

        # Hoja resumen
        ws_resumen.freeze_panes = 'A2'
        for cell in ws_resumen[1]:
            cell.font = header_font
            cell.fill = header_fill_resumen
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row in ws_resumen.iter_rows(min_row=2, max_row=ws_resumen.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

        ws_resumen.column_dimensions['A'].width = 36
        ws_resumen.column_dimensions['B'].width = 20

        # Formato porcentaje en métricas de ratio
        for fila in range(2, ws_resumen.max_row + 1):
            metrica = ws_resumen.cell(row=fila, column=1).value
            celda_valor = ws_resumen.cell(row=fila, column=2)
            if isinstance(metrica, str) and metrica.startswith('%'):
                celda_valor.number_format = '0.00%'


def procesar_excel(file_path):
    print(f"Cargando archivo {file_path}...")
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        print(f"Error al leer el excel: {e}")
        return

    # Verificar si la columna 'nombre' u otras similares existen
    col_nombre = None
    nombres_posibles = ['nombre', 'empresa', 'nombre_empresa', 'nombre empresa', 'razon social', 'razón social', 'company']
    for col in df.columns:
        if str(col).strip().lower() in nombres_posibles:
            col_nombre = col
            break

    if not col_nombre:
        print("El archivo Excel no contiene una columna llamada 'nombre'.")
        print(f"Columnas encontradas: {list(df.columns)}")
        return

    print("Archivo cargado correctamente. Iniciando enriquecimiento...")

    # Crear columnas si no existen
    if 'email' not in df.columns:
        df['email'] = None
    if 'gerente' not in df.columns:
        df['gerente'] = None

    total = len(df)
    for index, row in df.iterrows():
        nombre = row[col_nombre]
        if pd.isna(nombre) or not str(nombre).strip():
            continue

        nombre_str = str(nombre).strip()
        print(f"\n[{index+1}/{total}] Buscando datos para: {nombre_str}")

        lead_dict = {"nombre": nombre_str, "provincia": ""}
        resultado = enrich_lead(lead_dict)

        email = resultado.get("email")
        gerente = resultado.get("gerente")

        if email:
            df.at[index, 'email'] = email
        if gerente:
            df.at[index, 'gerente'] = gerente

        print(f"   -> Email: {email} | Gerente: {gerente}")
        time.sleep(1)  # Pequena pausa para no saturar

    # Guardar los resultados en un nuevo archivo
    output_file = file_path.replace('.xlsx', '_enriquecido.xlsx')
    if output_file == file_path:
        output_file = 'resultado_enriquecido.xlsx'

    try:
        _generar_excel_profesional(df, output_file, col_nombre)
        print(f"\nProceso completado con exito. Resultados guardados en: {output_file}")
    except Exception as e:
        print(f"Error al guardar el excel: {e}")


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        archivo = sys.argv[1]
    else:
        archivo = 'empresas.xlsx'
        print(f"No se especifico un archivo Excel como argumento. Buscando '{archivo}' por defecto...")

    if os.path.exists(archivo):
        procesar_excel(archivo)
    else:
        print(f"El archivo '{archivo}' no existe en la ruta actual.")
        print('Uso: python main.py <archivo.xlsx>')
