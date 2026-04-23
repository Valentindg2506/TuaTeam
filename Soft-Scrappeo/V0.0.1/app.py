"""
Servidor web — Radar Competidor
Ejecutar: python app.py  →  http://localhost:5000
Requiere: pip install flask requests openpyxl beautifulsoup4 lxml
"""

import io, re, time, threading, uuid
import requests as req
from bs4 import BeautifulSoup
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
jobs = {}

# ── Colores Excel ─────────────────────────────────────────────────────────────
C_TITLE = "062E1F"
C_HDR1  = "0A5C36"
C_HDR2  = "1A8754"
C_HDR3  = "2CB06E"
C_HDR4  = "3B4486"
C_EVEN  = "EBF7F1"
C_ODD   = "FFFFFF"
C_SUBE  = "D4EDDA"
C_BAJA  = "FDDEDE"
C_WHITE = "FFFFFF"
C_TEXT  = "1A1A1A"
C_LINK  = "0563C1"

def _border(color="C5D9CE", bottom_heavy=False):
    s = Side(style="thin", color=color)
    b = Side(style="medium", color="0A5C36") if bottom_heavy else s
    return Border(left=s, right=s, top=s, bottom=b)

THIN = _border()
HDRB = _border(bottom_heavy=True)

# ── URLs ──────────────────────────────────────────────────────────────────────
BASE_URL      = "https://ranking-empresas.eleconomista.es/ranking_empresas_nacional.html"
BASE_URL_PROV = "https://ranking-empresas.eleconomista.es/empresas-{p}.html"

PROV_URL = {
    "álava":"ALAVA","albacete":"ALBACETE","alicante":"ALICANTE","almería":"ALMERIA",
    "asturias":"ASTURIAS","ávila":"AVILA","badajoz":"BADAJOZ","barcelona":"BARCELONA",
    "bizkaia":"VIZCAYA","burgos":"BURGOS","cáceres":"CACERES","cádiz":"CADIZ",
    "cantabria":"CANTABRIA","castellon":"CASTELLON","ciudad real":"CIUDAD_REAL",
    "córdoba":"CORDOBA","coruña":"CORUNA","cuenca":"CUENCA","gipuzkoa":"GUIPUZCOA",
    "girona":"GERONA","granada":"GRANADA","guadalajara":"GUADALAJARA","huelva":"HUELVA",
    "huesca":"HUESCA","islas baleares":"BALEARES","jaén":"JAEN","la rioja":"RIOJA",
    "las palmas":"LAS_PALMAS","palmas (las)":"LAS_PALMAS","león":"LEON","lleida":"LERIDA",
    "lugo":"LUGO","madrid":"MADRID","málaga":"MALAGA","murcia":"MURCIA",
    "navarra":"NAVARRA","ourense":"ORENSE","palencia":"PALENCIA","pontevedra":"PONTEVEDRA",
    "salamanca":"SALAMANCA","segovia":"SEGOVIA","sevilla":"SEVILLA","soria":"SORIA",
    "tarragona":"TARRAGONA","tenerife":"TENERIFE","teruel":"TERUEL","toledo":"TOLEDO",
    "valencia":"VALENCIA","valladolid":"VALLADOLID","zamora":"ZAMORA","zaragoza":"ZARAGOZA",
}

PROVINCIAS = [
    "","Álava","Albacete","Alicante","Almería","Asturias","Ávila","Badajoz",
    "Barcelona","Bizkaia","Burgos","Cáceres","Cádiz","Cantabria","Castellon",
    "Ciudad Real","Córdoba","Coruña","Cuenca","Gipuzkoa","Girona","Granada",
    "Guadalajara","Huelva","Huesca","Islas Baleares","Jaén","La Rioja",
    "Las Palmas","León","Lleida","Lugo","Madrid","Málaga","Murcia","Navarra",
    "Ourense","Palencia","Palmas (las)","Pontevedra","Salamanca","Segovia",
    "Sevilla","Soria","Tarragona","Tenerife","Teruel","Toledo","Valencia",
    "Valladolid","Zamora","Zaragoza",
]

# ── Sesión HTTP ───────────────────────────────────────────────────────────────

def make_session():
    s = req.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Cache-Control": "max-age=0",
    })
    return s


def fetch_page(session, cnae, provincia, pagina):
    """
    Siempre usa la URL nacional con filtro de CNAE.
    El filtro de provincia se aplica localmente en parse_tabla.
    """
    url    = BASE_URL
    params = {"qSectorNorm": cnae}
    if pagina > 1:
        params["pagina"] = pagina

    for intento in range(3):
        try:
            if intento > 0:
                time.sleep(3 * intento)
            r = session.get(url, params=params, timeout=20)
            r.raise_for_status()
            return BeautifulSoup(r.text, "lxml"), None
        except req.HTTPError as e:
            code = e.response.status_code
            if code == 403:
                if intento < 2:
                    time.sleep(4 * (intento + 1))
                    continue
                return None, "Acceso bloqueado (403). Ejecuta la app desde tu PC local, no desde un servidor."
            if code == 404:
                return None, f"Provincia no encontrada: '{provincia}'."
            return None, f"Error HTTP {code}"
        except Exception as e:
            if intento < 2:
                time.sleep(2)
                continue
            return None, str(e)
    return None, "No se pudo conectar tras 3 intentos."

# ── Parseo ────────────────────────────────────────────────────────────────────

def parse_facturacion(texto):
    t = texto.strip()
    solo = t.replace(".", "").replace(",", "").replace(" ", "")
    if solo.isdigit() and len(solo) > 4:
        n = int(solo)
        return n, f"{n:,}".replace(",", ".")
    t_low = t.lower()
    for k, (num, label) in {
        "corporativa": (50_000_000, "Corporativa (>50M€)"),
        "grande":      (10_000_000, "Grande (>10M€)"),
        "mediana":     (2_000_000,  "Mediana (>2M€)"),
        "pequeña":     (500_000,    "Pequeña (<2M€)"),
        "pequena":     (500_000,    "Pequeña (<2M€)"),
    }.items():
        if k in t_low:
            return num, label
    return None, t or "—"


def parse_tabla(soup, cnae_filtro=None, provincia_forzada=None):
    rows = []
    table = soup.find("table")
    if not table: return rows
    for tr in table.find_all("tr"):
        tds = tr.find_all("td")
        n_cols = len(tds)
        # Saltar filas de cabecera (th) o filas con menos de 5 celdas de datos
        if n_cols < 5: continue

        pos_txt  = tds[0].get_text(strip=True)
        evol_txt = tds[1].get_text(strip=True)
        nombre   = tds[2].get_text(strip=True)
        factura  = tds[3].get_text(strip=True)
        sector   = tds[4].get_text(strip=True)
        # Col 5 = provincia (tabla nacional tiene 7 cols: pos,evol,nombre,fact,sector,prov,vermás)
        # Col 5 = "Ver más" (tabla provincial tiene 6 cols: pos,evol,nombre,fact,sector,vermás)
        if n_cols >= 7:
            prov = tds[5].get_text(strip=True)   # tabla nacional
        else:
            prov = provincia_forzada or ""         # tabla provincial

        if cnae_filtro and sector.strip() != str(cnae_filtro).strip():
            continue

        # Filtrar por provincia localmente si se especificó
        if provincia_forzada and prov.lower().strip() != provincia_forzada.lower().strip():
            continue

        link = ""
        for td in tds:
            a = td.find("a", href=True)
            if a:
                href = a["href"]
                if href and href not in ("#", "") and not href.startswith("javascript"):
                    link = href if href.startswith("http") else \
                           "https://ranking-empresas.eleconomista.es" + href
                    break

        try:    posicion = int(re.sub(r"\D", "", pos_txt))
        except: posicion = None

        evol_num, tendencia = 0, "Igual"
        m = re.search(r"(\d[\d\.]*)", evol_txt)
        if m:
            try: evol_num = int(m.group(1).replace(".", ""))
            except: pass
        if "Sube" in evol_txt:   tendencia = "Sube"
        elif "Baja" in evol_txt: tendencia = "Baja"

        factura_num, factura_label = parse_facturacion(factura)

        rows.append({
            "posicion_nacional":    posicion,
            "evolucion_posiciones": evol_num,
            "tendencia":            tendencia,
            "nombre":               nombre,
            "facturacion_raw":      factura_label,
            "facturacion_num":      factura_num,
            "cnae":                 sector,
            "provincia":            prov,
            "url_ficha":            link,
        })
    return rows


def top3_competidores(lead, pool, ratio_min, ratio_max):
    cands = [c for c in pool if c["nombre"] != lead["nombre"]]
    misma = [c for c in cands if c["provincia"] == lead["provincia"]]
    fuente = misma if misma else cands
    if lead["facturacion_num"] and lead["facturacion_num"] > 0:
        con_ratio = [c for c in fuente
                     if c["facturacion_num"] and
                        ratio_min <= c["facturacion_num"]/lead["facturacion_num"] <= ratio_max]
        if con_ratio: fuente = con_ratio
    return sorted([c for c in fuente if c["posicion_nacional"]],
                  key=lambda x: x["posicion_nacional"])[:3]

# ── Excel helpers ─────────────────────────────────────────────────────────────

def H(cell, text, bg, size=10):
    cell.value = text
    cell.font  = Font(name="Aptos Narrow", bold=True, color=C_WHITE, size=size)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = HDRB

def D(cell, value, bg=None, bold=False, align="left", num_fmt=None, link=None):
    cell.value = value
    cell.font  = Font(name="Aptos Narrow", size=10, bold=bold,
                      color=C_LINK if link else C_TEXT,
                      underline="single" if link else None)
    if bg: cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(vertical="center", horizontal=align)
    cell.border = THIN
    if num_fmt: cell.number_format = num_fmt
    if link and value: cell.hyperlink = value

def W(ws, widths):
    for c, w in widths.items():
        ws.column_dimensions[c].width = w

def row_bg(r): return C_EVEN if r % 2 == 0 else C_ODD

def title_row(ws, text, n_cols, row=1, h=32):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row, 1, text)
    c.font = Font(name="Aptos Narrow", bold=True, size=13, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=C_TITLE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = THIN
    ws.row_dimensions[row].height = h

# ── Hojas Excel ───────────────────────────────────────────────────────────────

def sheet_empresas(ws, datos, titulo, hdr_color, cnae, prov):
    tiene_tend = any(d["tendencia"] != "Igual" for d in datos)
    cols_base = ["Pos. Nacional","Empresa","Provincia","Facturación (€)","Evolución pos."]
    if tiene_tend: cols_base.insert(4, "Tendencia")
    cols_base.append("Ficha (enlace)")
    n = len(cols_base)
    loc = prov or "Nacional"
    title_row(ws, f"{titulo} — CNAE {cnae} · {loc}", n)
    ws.row_dimensions[2].height = 28
    for c, h in enumerate(cols_base, 1): H(ws.cell(2, c), h, hdr_color)
    for r, row in enumerate(datos, 3):
        bg = row_bg(r)
        if tiene_tend:
            if row["tendencia"] == "Sube": bg = C_SUBE
            elif row["tendencia"] == "Baja": bg = C_BAJA
        fval = row["facturacion_num"] if (row["facturacion_num"] and
               not any(x in row["facturacion_raw"] for x in
                       ("Corporativa","Grande","Mediana","Pequeña"))) \
               else row["facturacion_raw"]
        fnf = "#,##0" if isinstance(fval, int) else None
        col = 1
        D(ws.cell(r, col), row["posicion_nacional"], bg, align="center", num_fmt="#,##0"); col+=1
        D(ws.cell(r, col), row["nombre"], bg); col+=1
        D(ws.cell(r, col), row["provincia"], bg); col+=1
        if tiene_tend: D(ws.cell(r, col), row["tendencia"], bg, align="center"); col+=1
        D(ws.cell(r, col), fval, bg, align="right", num_fmt=fnf); col+=1
        D(ws.cell(r, col), row["evolucion_posiciones"], bg, align="center", num_fmt="#,##0"); col+=1
        D(ws.cell(r, col), row["url_ficha"] or "—", bg, link=bool(row["url_ficha"]))
        ws.row_dimensions[r].height = 16
    w = {"A":13,"B":42,"C":14,"D":20,"E":14,"F":55}
    if tiene_tend: w = {"A":13,"B":42,"C":14,"D":12,"E":20,"F":14,"G":55}
    W(ws, w)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}2"


def sheet_resumen(ws, resumen, cnae, prov):
    max_comp = min(max((len(r["competidores"]) for r in resumen), default=1), 3)
    lead_w, comp_w = 4, 4
    n_cols = lead_w + comp_w * max_comp
    loc = prov or "Nacional"
    title_row(ws, f"Análisis Lead vs. Competidores — CNAE {cnae} · {loc}", n_cols)
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:D2")
    gc = ws.cell(2, 1, "LEAD")
    gc.font = Font(name="Aptos Narrow", bold=True, size=9, color=C_WHITE)
    gc.fill = PatternFill("solid", start_color=C_HDR1)
    gc.alignment = Alignment(horizontal="center", vertical="center")
    gc.border = THIN
    comp_colors = [C_HDR2, C_HDR3, C_HDR4]
    for i in range(max_comp):
        cs = lead_w + i*comp_w + 1
        ce = cs + comp_w - 1
        ws.merge_cells(f"{get_column_letter(cs)}2:{get_column_letter(ce)}2")
        gc2 = ws.cell(2, cs, f"COMPETIDOR {i+1}")
        gc2.font = Font(name="Aptos Narrow", bold=True, size=9, color=C_WHITE)
        gc2.fill = PatternFill("solid", start_color=comp_colors[i])
        gc2.alignment = Alignment(horizontal="center", vertical="center")
        gc2.border = THIN
    ws.row_dimensions[3].height = 28
    for c, h in enumerate(["Empresa (Lead)","Provincia","Facturación","Pos. Nacional"], 1):
        H(ws.cell(3, c), h, C_HDR1, size=9)
    for i in range(max_comp):
        cs = lead_w + i*comp_w + 1
        for j, h in enumerate(["Competidor","Provincia","Facturación","Ratio ×"]):
            H(ws.cell(3, cs+j), h, comp_colors[i], size=9)
    for r, row in enumerate(resumen, 4):
        bg = row_bg(r)
        fval = row["lead_facturacion_num"] if (row["lead_facturacion_num"] and
               not any(x in row["lead_facturacion"] for x in
                       ("Corporativa","Grande","Mediana","Pequeña"))) \
               else row["lead_facturacion"]
        fnf = "#,##0" if isinstance(fval, int) else None
        D(ws.cell(r,1), row["lead_nombre"], bg)
        D(ws.cell(r,2), row["lead_provincia"], bg)
        D(ws.cell(r,3), fval, bg, align="right", num_fmt=fnf)
        D(ws.cell(r,4), row["lead_posicion"], bg, align="center", num_fmt="#,##0")
        for i, comp in enumerate(row["competidores"][:max_comp]):
            cs  = lead_w + i*comp_w + 1
            cbg = C_SUBE if comp.get("mismo_area") == "Sí" else bg
            cf  = comp["facturacion_num"] if (comp["facturacion_num"] and
                  not any(x in comp["facturacion_raw"] for x in
                          ("Corporativa","Grande","Mediana","Pequeña"))) \
                  else comp["facturacion_raw"]
            cfn = "#,##0" if isinstance(cf, int) else None
            D(ws.cell(r, cs),   comp["nombre"],    cbg)
            D(ws.cell(r, cs+1), comp["provincia"], cbg)
            D(ws.cell(r, cs+2), cf, cbg, align="right", num_fmt=cfn)
            D(ws.cell(r, cs+3), comp["ratio"], cbg, align="center",
              bold=True, num_fmt='0.0"×"')
        for i in range(len(row["competidores"]), max_comp):
            cs = lead_w + i*comp_w + 1
            for j in range(comp_w): D(ws.cell(r, cs+j), "—", C_EVEN)
        ws.row_dimensions[r].height = 16
    widths = {"A":38,"B":14,"C":18,"D":13}
    for i in range(max_comp):
        for j, w in enumerate([36, 14, 18, 9]):
            widths[get_column_letter(lead_w + i*comp_w + 1 + j)] = w
    W(ws, widths)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"


def sheet_leyenda(ws):
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 68
    items = [
        ("ESTRUCTURA DEL EXCEL", None),
        ("Hoja 1 — Leads",               "Lista de empresas medianas/pequeñas del sector: tus posibles clientes."),
        ("Hoja 2 — Lead vs Competidores", "Para cada lead, sus 3 competidores más relevantes del mismo sector y zona."),
        ("Hoja 3 — Leyenda",             "Esta hoja: explicación de campos y colores."),
        ("",""),
        ("CAMPOS CLAVE", None),
        ("Ratio ×",       "Facturación del competidor ÷ facturación del lead. Rango óptimo: 3×–20×."),
        ("Verde claro",   "Competidor en la misma provincia que el lead."),
        ("Blanco/gris",   "Competidor de otra provincia (válido como referencia sectorial)."),
        ("",""),
        ("FACTURACIÓN", None),
        ("Cifra exacta",        "Dato publicado en el Registro Mercantil / INFORMA D&B."),
        ("Corporativa (>50M€)", "Empresa que no publica cifras exactas."),
        ("Grande (>10M€)",      "Rango estimado por eInforma."),
        ("Mediana (>2M€)",      "Rango estimado por eInforma."),
        ("Pequeña (<2M€)",      "Rango estimado por eInforma."),
        ("",""),
        ("FUENTE", None),
        ("Sitio",     "ranking-empresas.eleconomista.es"),
        ("Proveedor", "eInforma / INFORMA D&B S.A.U."),
        ("Ejercicio", "Datos fiscales 2023, publicados en 2024"),
    ]
    for r, (k, v) in enumerate(items, 1):
        ws.row_dimensions[r].height = 18
        if v is None:
            ws.merge_cells(f"A{r}:B{r}")
            c = ws.cell(r, 1, k)
            c.font = Font(name="Aptos Narrow", bold=True, color=C_WHITE, size=10)
            c.fill = PatternFill("solid", start_color=C_HDR1)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = THIN
        elif k == "":
            pass
        else:
            c1 = ws.cell(r, 1, k)
            c1.font = Font(name="Aptos Narrow", bold=True, size=10, color=C_TEXT)
            c1.fill = PatternFill("solid", start_color=C_EVEN)
            c1.border = THIN
            c2 = ws.cell(r, 2, v)
            c2.font = Font(name="Aptos Narrow", size=10, color=C_TEXT)
            c2.fill = PatternFill("solid", start_color=C_ODD)
            c2.border = THIN


def build_excel(leads, competidores, resumen, cnae, prov):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Leads"
    ws1.sheet_properties.tabColor = "0A5C36"
    sheet_empresas(ws1, leads, "Leads", C_HDR1, cnae, prov)
    ws2 = wb.create_sheet("Lead vs Competidores")
    ws2.sheet_properties.tabColor = "062E1F"
    sheet_resumen(ws2, resumen, cnae, prov)
    ws3 = wb.create_sheet("Leyenda")
    ws3.sheet_properties.tabColor = "888888"
    sheet_leyenda(ws3)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Worker thread ─────────────────────────────────────────────────────────────

def run_job(job_id, cnae, prov, n_comp, n_leads, ratio_min, ratio_max, delay):
    def upd(status, pct, msg):
        jobs[job_id].update({"status": status, "progress": pct, "message": msg})
    try:
        session = make_session()
        upd("running", 5, "Conectando…")

        # Warm-up: visitar la home para obtener cookies reales
        try:
            session.get("https://ranking-empresas.eleconomista.es/", timeout=15)
            time.sleep(2)
        except: pass

        # La URL siempre es nacional con filtro CNAE.
        # La provincia se filtra localmente en parse_tabla.
        prov_filtro = prov if prov else None

        competidores = []
        for p in range(1, n_comp + 1):
            upd("running", 5 + int(p/(n_comp+n_leads)*65),
                f"Competidores — página {p}/{n_comp}…")
            soup, err = fetch_page(session, cnae, None, p)
            if err: upd("error", 0, err); return
            rows = parse_tabla(soup, provincia_forzada=prov_filtro)
            if not rows and p == 1: break
            competidores.extend(rows)
            time.sleep(delay)

        upd("running", 40, f"{len(competidores)} competidores. Descargando leads…")
        leads = []
        for i, p in enumerate(range(n_comp+1, n_comp+n_leads+1)):
            upd("running", 40+int(i/n_leads*40), f"Leads — página {p}…")
            soup, err = fetch_page(session, cnae, None, p)
            if err: upd("error", 0, err); return
            rows = parse_tabla(soup, provincia_forzada=prov_filtro)
            leads.extend(rows)
            time.sleep(delay)

        if not leads and not competidores:
            upd("error", 0, "Sin datos. Comprueba el CNAE — puede que no haya empresas de ese sector en esa provincia.")
            return

        upd("running", 85, f"Cruzando {len(leads)} leads…")
        resumen = []
        for lead in leads:
            top3 = top3_competidores(lead, competidores, ratio_min, ratio_max)
            comp_list = []
            for c in top3:
                ratio = round(c["facturacion_num"]/lead["facturacion_num"], 1) \
                        if (c["facturacion_num"] and lead["facturacion_num"]) else None
                comp_list.append({**c, "ratio": ratio,
                                   "mismo_area": "Sí" if c["provincia"]==lead["provincia"] else "No"})
            resumen.append({
                "lead_nombre":          lead["nombre"],
                "lead_provincia":       lead["provincia"],
                "lead_facturacion":     lead["facturacion_raw"],
                "lead_facturacion_num": lead["facturacion_num"],
                "lead_posicion":        lead["posicion_nacional"],
                "competidores":         comp_list,
            })

        upd("running", 95, "Generando Excel…")
        buf = build_excel(leads, competidores, resumen, cnae, prov)
        jobs[job_id]["file"] = buf
        jobs[job_id]["filename"] = f"radar_{cnae}{'_'+prov.replace(' ','_') if prov else ''}.xlsx"
        upd("done", 100, f"✅ {len(leads)} leads · {len(competidores)} competidores")

    except Exception as e:
        upd("error", 0, f"Error: {e}")

# ── Flask ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html", provincias=PROVINCIAS)

@app.route("/test")
def test():
    """Diagnóstico — abre http://localhost:5000/test"""
    import traceback
    resultado = {}
    try:
        session = make_session()
        session.get("https://ranking-empresas.eleconomista.es/", timeout=15)
        time.sleep(1)
        r = session.get(BASE_URL, params={"qSectorNorm": "4662"}, timeout=20)
        resultado["status_code"] = r.status_code
        soup = BeautifulSoup(r.text, "lxml")
        table = soup.find("table")
        if table:
            filas = table.find_all("tr")
            resultado["total_filas"] = len(filas)
            # Mostrar las primeras 5 filas con detalle de columnas
            muestra = []
            for i, tr in enumerate(filas[:6]):
                tds = tr.find_all("td")
                ths = tr.find_all("th")
                muestra.append({
                    "fila": i,
                    "num_td": len(tds),
                    "num_th": len(ths),
                    "texto": tr.get_text("|", strip=True)[:200],
                })
            resultado["muestra_filas"] = muestra
    except Exception as e:
        resultado["error"] = str(e)
        resultado["traceback"] = traceback.format_exc()
    return jsonify(resultado)

@app.route("/start", methods=["POST"])
def start():
    d = request.json
    cnae = d.get("cnae","").strip()
    if not cnae: return jsonify({"error":"El CNAE es obligatorio"}), 400
    jid = str(uuid.uuid4())
    jobs[jid] = {"status":"pending","progress":0,"message":"En cola…","file":None}
    threading.Thread(target=run_job, daemon=True, args=(
        jid, cnae, d.get("provincia","").strip(),
        max(1,min(int(d.get("n_comp",2)),5)),
        max(1,min(int(d.get("n_leads",3)),10)),
        float(d.get("ratio_min",3)), float(d.get("ratio_max",20)),
        float(d.get("delay",2.0)),
    )).start()
    return jsonify({"job_id": jid})

@app.route("/status/<jid>")
def status(jid):
    j = jobs.get(jid)
    if not j: return jsonify({"error":"no encontrado"}), 404
    return jsonify({"status":j["status"],"progress":j["progress"],
                    "message":j["message"],"ready":j["status"]=="done"})

@app.route("/download/<jid>")
def download(jid):
    j = jobs.get(jid)
    if not j or j["status"]!="done" or not j["file"]:
        return "Archivo no disponible", 404
    j["file"].seek(0)
    return send_file(j["file"], as_attachment=True,
                     download_name=j.get("filename","radar.xlsx"),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    print("\n" + "="*50)
    print("  Radar Competidor — http://localhost:5000")
    print("="*50 + "\n")
    app.run(debug=False, port=5000)
