"""
Radar Competidor — Servidor Flask
Ejecutar:  python app.py  →  http://localhost:5000

Dependencias:
    pip install flask requests beautifulsoup4 lxml openpyxl
"""

import io
import re
import time
import threading
import uuid
from urllib.parse import unquote

import requests as req
from bs4 import BeautifulSoup
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

app = Flask(__name__)
jobs = {}  # job_id → {status, progress, message, file, filename}

# ═════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═════════════════════════════════════════════════════════════════════════════

BASE_URL = "https://ranking-empresas.eleconomista.es/ranking_empresas_nacional.html"
EMPRESAS_PER_PAGE = 100  # el sitio muestra 100 empresas por página

# Provincias oficiales de España (orden alfabético, como aparecen en el sitio)
PROVINCIAS = [
    "", "Álava", "Albacete", "Alicante", "Almería", "Asturias", "Ávila",
    "Badajoz", "Barcelona", "Bizkaia", "Burgos", "Cáceres", "Cádiz",
    "Cantabria", "Castellón", "Ciudad Real", "Córdoba", "Coruña", "Cuenca",
    "Gerona", "Gipuzkoa", "Granada", "Guadalajara", "Huelva", "Huesca",
    "Islas Baleares", "Jaén", "La Rioja", "Las Palmas", "León", "Lleida",
    "Lugo", "Madrid", "Málaga", "Murcia", "Navarra", "Ourense", "Palencia",
    "Palmas (las)", "Pontevedra", "Salamanca", "Segovia", "Sevilla", "Soria",
    "Tarragona", "Tenerife", "Teruel", "Toledo", "Valencia", "Valladolid",
    "Zamora", "Zaragoza",
]

# Normalización de nombres de provincia (para comparar sin acentos/mayúsculas)
def normalizar(txt):
    if not txt: return ""
    txt = txt.lower().strip()
    reemplazos = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n",
        "à": "a", "è": "e", "ì": "i", "ò": "o", "ù": "u",
    }
    for k, v in reemplazos.items():
        txt = txt.replace(k, v)
    return txt

# ═════════════════════════════════════════════════════════════════════════════
# SCRAPING
# ═════════════════════════════════════════════════════════════════════════════

def make_session():
    s = req.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "es-ES,es;q=0.9",
        "Connection": "keep-alive",
    })
    return s


def fetch_page(session, cnae, pagina):
    """Descarga una página del ranking nacional filtrado por CNAE."""
    params = {"qSectorNorm": cnae}
    if pagina > 1:
        params["pagina"] = pagina

    for intento in range(3):
        try:
            if intento > 0:
                time.sleep(2 * intento)
            r = session.get(BASE_URL, params=params, timeout=20)
            r.raise_for_status()
            return BeautifulSoup(r.text, "lxml"), None
        except req.HTTPError as e:
            code = e.response.status_code
            if code == 403 and intento < 2:
                time.sleep(4)
                continue
            return None, f"Error HTTP {code}"
        except Exception as e:
            if intento < 2:
                time.sleep(2)
                continue
            return None, str(e)
    return None, "No se pudo conectar tras 3 intentos."


def parse_facturacion(texto):
    """
    El sitio mezcla cifras reales ('167.331.470') con rangos ('corporativa',
    'grande', 'mediana', 'pequeña') para empresas que no publican su cifra.
    Devuelve (valor_numerico_estimado, etiqueta_legible).
    """
    t = (texto or "").strip()
    solo = t.replace(".", "").replace(",", "").replace(" ", "")
    if solo.isdigit() and len(solo) > 4:
        n = int(solo)
        return n, f"{n:,.0f}".replace(",", ".") + " €"
    t_low = t.lower()
    rangos = [
        ("corporativa", 50_000_000, "Corporativa (>50M€)"),
        ("grande",      10_000_000, "Grande (>10M€)"),
        ("mediana",      2_000_000, "Mediana (>2M€)"),
        ("pequeña",        500_000, "Pequeña (<2M€)"),
        ("pequena",        500_000, "Pequeña (<2M€)"),
    ]
    for clave, num, label in rangos:
        if clave in t_low:
            return num, label
    return None, t or "—"


def parse_tabla(soup, provincia_filtro=None):
    """
    Parsea la tabla principal del ranking. Cada fila tiene 7 columnas:
    posición | evolución | nombre (link) | facturación | CNAE | provincia | botón
    """
    rows = []
    table = soup.find("table")
    if not table:
        return rows

    for tr in table.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 7:
            continue  # fila de cabecera o filtros

        pos_txt  = tds[0].get_text(strip=True)
        evol_txt = tds[1].get_text(strip=True)
        nombre   = tds[2].get_text(strip=True)
        factura  = tds[3].get_text(strip=True)
        sector   = tds[4].get_text(strip=True)
        prov     = tds[5].get_text(strip=True)

        # Filtrar por provincia si se especificó
        if provincia_filtro:
            if normalizar(prov) != normalizar(provincia_filtro):
                continue

        # URL de la ficha (enlace en la columna del nombre o "Ver más")
        url = ""
        for td in tds:
            a = td.find("a", href=True)
            if a:
                href = a["href"]
                if href and not href.startswith(("javascript:", "#")):
                    url = href if href.startswith("http") else \
                          "https://ranking-empresas.eleconomista.es" + href
                    break

        # Parsear posición
        try:
            posicion = int(re.sub(r"\D", "", pos_txt))
        except ValueError:
            posicion = None

        # Parsear evolución + tendencia
        evol_num = None
        tendencia = "Igual"
        m = re.search(r"([\d\.]+)", evol_txt)
        if m:
            try:
                evol_num = int(m.group(1).replace(".", ""))
            except ValueError:
                pass
        if "Sube" in evol_txt:
            tendencia = "Sube"
        elif "Baja" in evol_txt:
            tendencia = "Baja"
        elif "(ND)" in evol_txt or "ND" in evol_txt:
            tendencia = "ND"

        fact_num, fact_label = parse_facturacion(factura)

        rows.append({
            "posicion":     posicion,
            "evolucion":    evol_num,
            "tendencia":    tendencia,
            "nombre":       nombre,
            "fact_num":     fact_num,
            "fact_label":   fact_label,
            "cnae":         sector,
            "provincia":    prov,
            "url":          url,
        })
    return rows


def encontrar_competidores(lead, pool, ratio_min, ratio_max, n=3):
    """
    Devuelve los N mejores competidores para un lead.
    Prioriza: misma provincia → ratio tamaño 3-20× → mejor posición.
    """
    candidatos = [c for c in pool if c["nombre"] != lead["nombre"]]
    if not candidatos:
        return []

    # Paso 1: priorizar misma provincia
    misma_prov = [c for c in candidatos if c["provincia"] == lead["provincia"]]
    fuente = misma_prov if len(misma_prov) >= n else candidatos

    # Paso 2: filtrar por ratio de tamaño si hay datos
    if lead["fact_num"]:
        con_ratio = [
            c for c in fuente
            if c["fact_num"] and ratio_min <= c["fact_num"]/lead["fact_num"] <= ratio_max
        ]
        if len(con_ratio) >= n:
            fuente = con_ratio

    # Paso 3: ordenar por posición (los mejores del ranking primero)
    ordenados = sorted(
        [c for c in fuente if c["posicion"]],
        key=lambda x: x["posicion"]
    )
    return ordenados[:n]


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL
# ═════════════════════════════════════════════════════════════════════════════

# Paleta profesional verde
C_TITLE    = "0D4225"
C_HEADER   = "146C43"
C_HEADER2  = "1E8F5A"
C_HEADER3  = "2AAE70"
C_ACCENT   = "C89B3C"   # dorado suave
C_BG_LIGHT = "F0F8F4"
C_BG_MED   = "DEF0E5"
C_SUBE     = "D4EDDA"
C_BAJA     = "F8D7DA"
C_WHITE    = "FFFFFF"
C_TEXT     = "1B1B1B"
C_TEXT_MUT = "5E6C67"
C_LINK     = "0D6EFD"

FONT_NAME = "Aptos Narrow"

def _side(color="B7D3C2", style="thin"):
    return Side(style=style, color=color)

def _border(bottom_heavy=False, top_heavy=False):
    return Border(
        left   = _side(),
        right  = _side(),
        top    = _side(color="0D4225", style="medium") if top_heavy else _side(),
        bottom = _side(color="0D4225", style="medium") if bottom_heavy else _side(),
    )


def _H(cell, text, bg=C_HEADER, size=10, align="center"):
    cell.value = text
    cell.font = Font(name=FONT_NAME, bold=True, color=C_WHITE, size=size)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _border(bottom_heavy=True)

def _D(cell, value, bg=None, align="left", num_fmt=None, bold=False,
       hyperlink=None, color=C_TEXT, size=10):
    cell.value = value
    cell.font = Font(
        name=FONT_NAME, size=size, bold=bold,
        color=C_LINK if hyperlink else color,
        underline="single" if hyperlink else None,
    )
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        vertical="center", horizontal=align, wrap_text=False
    )
    cell.border = _border()
    if num_fmt:
        cell.number_format = num_fmt
    if hyperlink:
        cell.hyperlink = hyperlink

def _set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _row_bg(r):
    """Bandas alternas — filas pares gris claro, impares blanco."""
    return C_BG_LIGHT if r % 2 == 0 else C_WHITE


def _title_block(ws, titulo, subtitulo, n_cols):
    """Bloque de título + subtítulo en las 2 primeras filas."""
    # Fila 1: título
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws.cell(1, 1, titulo)
    c.font = Font(name=FONT_NAME, bold=True, size=14, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=C_TITLE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # Fila 2: subtítulo
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    c2 = ws.cell(2, 1, subtitulo)
    c2.font = Font(name=FONT_NAME, italic=True, size=10, color=C_TEXT_MUT)
    c2.fill = PatternFill("solid", start_color=C_BG_LIGHT)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # Fila 3: separador fino
    ws.row_dimensions[3].height = 6


# ── Hoja 1: Leads ────────────────────────────────────────────────────────────

def sheet_leads(ws, leads, cnae, provincia):
    loc = provincia if provincia else "Nacional"
    _title_block(
        ws,
        f"LEADS — Sector CNAE {cnae}",
        f"Ámbito: {loc} · {len(leads)} empresas encontradas",
        7
    )

    headers = [
        "Pos. Ranking", "Empresa", "Provincia", "Facturación",
        "Evolución (±)", "Tendencia", "Ficha"
    ]
    for c, h in enumerate(headers, 1):
        _H(ws.cell(4, c), h, bg=C_HEADER)
    ws.row_dimensions[4].height = 30

    for r, lead in enumerate(leads, 5):
        bg = _row_bg(r)
        _D(ws.cell(r, 1), lead["posicion"], bg, align="center", num_fmt="#,##0")
        _D(ws.cell(r, 2), lead["nombre"], bg, bold=True)
        _D(ws.cell(r, 3), lead["provincia"], bg)

        if isinstance(lead["fact_num"], int) and "(" not in lead["fact_label"]:
            _D(ws.cell(r, 4), lead["fact_num"], bg, align="right", num_fmt="#,##0 €")
        else:
            _D(ws.cell(r, 4), lead["fact_label"], bg, align="right",
               color=C_TEXT_MUT, size=9)

        _D(ws.cell(r, 5), lead["evolucion"], bg, align="center", num_fmt="#,##0")

        # Tendencia con emoji + color
        t = lead["tendencia"]
        tend_bg = C_SUBE if t == "Sube" else (C_BAJA if t == "Baja" else bg)
        tend_txt = {"Sube": "▲ Sube", "Baja": "▼ Baja",
                    "Igual": "● Igual", "ND": "— ND"}.get(t, "— ND")
        _D(ws.cell(r, 6), tend_txt, tend_bg, align="center", bold=(t in ("Sube","Baja")))

        if lead["url"]:
            _D(ws.cell(r, 7), "Ver ficha →", bg, align="center",
               hyperlink=lead["url"], size=9)
        else:
            _D(ws.cell(r, 7), "—", bg, align="center", color=C_TEXT_MUT)

        ws.row_dimensions[r].height = 18

    _set_widths(ws, {"A":13, "B":45, "C":15, "D":18, "E":14, "F":14, "G":14})
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{4 + len(leads)}"
    # print area
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.print_title_rows = "1:4"


# ── Hoja 2: Lead vs Competidores ─────────────────────────────────────────────

def sheet_resumen(ws, resumen, cnae, provincia, max_comp=3):
    loc = provincia if provincia else "Nacional"
    n_cols = 5 + 4 * max_comp
    _title_block(
        ws,
        f"ANÁLISIS LEAD vs. COMPETIDORES — Sector CNAE {cnae}",
        f"Ámbito: {loc} · {len(resumen)} leads · hasta {max_comp} competidores por lead",
        n_cols
    )

    # Fila 4: super-cabeceras de grupo
    ws.row_dimensions[4].height = 22

    ws.merge_cells(f"A4:E4")
    gc = ws.cell(4, 1, "LEAD (EMPRESA OBJETIVO)")
    gc.font = Font(name=FONT_NAME, bold=True, size=10, color=C_WHITE)
    gc.fill = PatternFill("solid", start_color=C_HEADER)
    gc.alignment = Alignment(horizontal="center", vertical="center")
    gc.border = _border()

    comp_colors = [C_HEADER2, C_HEADER3, "3BC47D"]
    for i in range(max_comp):
        cs = 6 + i * 4
        ce = cs + 3
        ws.merge_cells(f"{get_column_letter(cs)}4:{get_column_letter(ce)}4")
        gc2 = ws.cell(4, cs, f"COMPETIDOR {i+1}")
        gc2.font = Font(name=FONT_NAME, bold=True, size=10, color=C_WHITE)
        gc2.fill = PatternFill("solid", start_color=comp_colors[i])
        gc2.alignment = Alignment(horizontal="center", vertical="center")
        gc2.border = _border()

    # Fila 5: cabeceras de campo
    ws.row_dimensions[5].height = 30
    lead_hdrs = ["Empresa", "Provincia", "Facturación", "Pos.", "Tendencia"]
    for c, h in enumerate(lead_hdrs, 1):
        _H(ws.cell(5, c), h, bg=C_HEADER, size=9)

    comp_hdrs = ["Empresa Competidora", "Provincia", "Facturación", "Ratio ×"]
    for i in range(max_comp):
        cs = 6 + i * 4
        for j, h in enumerate(comp_hdrs):
            _H(ws.cell(5, cs + j), h, bg=comp_colors[i], size=9)

    # Filas de datos
    for r_idx, row in enumerate(resumen, 6):
        bg = _row_bg(r_idx)
        lead = row["lead"]

        _D(ws.cell(r_idx, 1), lead["nombre"], bg, bold=True)
        _D(ws.cell(r_idx, 2), lead["provincia"], bg)
        if isinstance(lead["fact_num"], int) and "(" not in lead["fact_label"]:
            _D(ws.cell(r_idx, 3), lead["fact_num"], bg, align="right", num_fmt="#,##0 €")
        else:
            _D(ws.cell(r_idx, 3), lead["fact_label"], bg, align="right",
               color=C_TEXT_MUT, size=9)
        _D(ws.cell(r_idx, 4), lead["posicion"], bg, align="center", num_fmt="#,##0")

        t = lead["tendencia"]
        tend_txt = {"Sube": "▲", "Baja": "▼",
                    "Igual": "●", "ND": "—"}.get(t, "—")
        tend_color = "2D8659" if t == "Sube" else ("C0392B" if t == "Baja" else C_TEXT_MUT)
        _D(ws.cell(r_idx, 5), tend_txt, bg, align="center", bold=True, color=tend_color)

        # Competidores
        for i, comp in enumerate(row["competidores"][:max_comp]):
            cs = 6 + i * 4
            # Fondo verde claro si comparten provincia (prueba social local)
            cbg = C_BG_MED if comp["provincia"] == lead["provincia"] else bg

            _D(ws.cell(r_idx, cs), comp["nombre"], cbg, bold=True)
            _D(ws.cell(r_idx, cs+1), comp["provincia"], cbg)

            if isinstance(comp["fact_num"], int) and "(" not in comp["fact_label"]:
                _D(ws.cell(r_idx, cs+2), comp["fact_num"], cbg,
                   align="right", num_fmt="#,##0 €")
            else:
                _D(ws.cell(r_idx, cs+2), comp["fact_label"], cbg,
                   align="right", color=C_TEXT_MUT, size=9)

            _D(ws.cell(r_idx, cs+3), comp.get("ratio"), cbg,
               align="center", bold=True, num_fmt='0.0"×"')

        # Rellenar huecos cuando hay <max_comp competidores
        for i in range(len(row["competidores"]), max_comp):
            cs = 6 + i * 4
            for j in range(4):
                _D(ws.cell(r_idx, cs+j), "—", C_BG_LIGHT,
                   align="center", color=C_TEXT_MUT)

        ws.row_dimensions[r_idx].height = 20

    widths = {"A": 38, "B": 14, "C": 17, "D": 9, "E": 11}
    for i in range(max_comp):
        for j, w in enumerate([36, 14, 17, 9]):
            widths[get_column_letter(6 + i*4 + j)] = w
    _set_widths(ws, widths)
    ws.freeze_panes = "F6"
    ws.auto_filter.ref = f"A5:{get_column_letter(n_cols)}{5 + len(resumen)}"
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.print_title_rows = "1:5"


# ── Hoja 3: Leyenda ──────────────────────────────────────────────────────────

def sheet_leyenda(ws):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 75

    # Título
    ws.merge_cells("A1:B1")
    c = ws.cell(1, 1, "📖 GUÍA DE USO — RADAR COMPETIDOR")
    c.font = Font(name=FONT_NAME, bold=True, size=14, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=C_TITLE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    secciones = [
        ("ESTRUCTURA DEL EXCEL", None),
        ("Hoja 1 — Leads",
         "Empresas medianas del sector CNAE buscado, ordenadas por posición "
         "en el ranking nacional. Son tus potenciales clientes."),
        ("Hoja 2 — Lead vs Competidores",
         "Para cada lead, hasta 3 competidores relevantes del mismo sector. "
         "Úsalos como prueba social al contactar: \"Ayudamos a [competidor] "
         "a ganar licitaciones — podemos hacer lo mismo por vosotros\"."),
        ("Hoja 3 — Leyenda",
         "Esta hoja."),

        ("", ""),
        ("CAMPOS CLAVE", None),
        ("Pos. Ranking",
         "Posición nacional por facturación (de 1 a 500.000). Menor = más grande."),
        ("Facturación",
         "Ventas del ejercicio 2023. Puede ser una cifra exacta (Registro "
         "Mercantil) o un rango estimado si la empresa no publica datos."),
        ("Evolución (±)",
         "Número de posiciones que la empresa ha subido o bajado respecto al "
         "año anterior en el ranking."),
        ("Tendencia",
         "▲ Sube = crecimiento · ▼ Baja = pérdida de facturación · ● Igual = sin cambio."),
        ("Ratio ×",
         "Cuántas veces más grande es el competidor que el lead "
         "(facturación competidor ÷ facturación lead). "
         "Rango útil para prueba social: 3× a 20×."),

        ("", ""),
        ("CÓMO LEER LOS COLORES", None),
        ("Verde claro (en Hoja 2)",
         "El competidor está en la misma provincia que el lead → mayor "
         "reconocimiento, es muy probable que el lead ya conozca a ese competidor."),
        ("Blanco (en Hoja 2)",
         "Competidor de otra provincia — válido como referencia sectorial "
         "aunque menos impactante que uno local."),

        ("", ""),
        ("TIPOS DE FACTURACIÓN", None),
        ("Cifra exacta (ej. 45.000.000 €)",
         "Publicada en el Registro Mercantil. Fiable."),
        ("Corporativa (>50M€)",
         "Empresa grande que no depositó cuentas públicas. Facturación real probablemente muy alta."),
        ("Grande (>10M€)",  "Rango estimado por eInforma."),
        ("Mediana (>2M€)",  "Rango estimado por eInforma."),
        ("Pequeña (<2M€)",  "Rango estimado por eInforma."),

        ("", ""),
        ("FUENTE", None),
        ("Web",       "ranking-empresas.eleconomista.es"),
        ("Proveedor", "eInforma / INFORMA D&B S.A.U."),
        ("Ejercicio", "Datos fiscales del año 2023 (publicados en 2024)."),
    ]

    for r, (k, v) in enumerate(secciones, 2):
        ws.row_dimensions[r].height = 22 if v is None else 32

        if v is None:  # cabecera de sección
            ws.merge_cells(f"A{r}:B{r}")
            c = ws.cell(r, 1, k)
            c.font = Font(name=FONT_NAME, bold=True, size=11, color=C_WHITE)
            c.fill = PatternFill("solid", start_color=C_HEADER)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = _border()
        elif k == "":  # línea vacía
            ws.row_dimensions[r].height = 8
        else:
            c1 = ws.cell(r, 1, k)
            c1.font = Font(name=FONT_NAME, bold=True, size=10, color=C_TEXT)
            c1.fill = PatternFill("solid", start_color=C_BG_LIGHT)
            c1.alignment = Alignment(
                horizontal="left", vertical="center", indent=1, wrap_text=True
            )
            c1.border = _border()

            c2 = ws.cell(r, 2, v)
            c2.font = Font(name=FONT_NAME, size=10, color=C_TEXT)
            c2.fill = PatternFill("solid", start_color=C_WHITE)
            c2.alignment = Alignment(
                horizontal="left", vertical="center", indent=1, wrap_text=True
            )
            c2.border = _border()


def build_excel(leads, resumen, cnae, provincia, max_comp=3):
    wb = Workbook()
    wb.properties.creator = "Radar Competidor"
    wb.properties.title   = f"Análisis CNAE {cnae}"

    ws1 = wb.active
    ws1.title = "Leads"
    ws1.sheet_properties.tabColor = C_HEADER
    sheet_leads(ws1, leads, cnae, provincia)

    ws2 = wb.create_sheet("Lead vs Competidores")
    ws2.sheet_properties.tabColor = C_TITLE
    sheet_resumen(ws2, resumen, cnae, provincia, max_comp)

    ws3 = wb.create_sheet("Leyenda")
    ws3.sheet_properties.tabColor = "808080"
    sheet_leyenda(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═════════════════════════════════════════════════════════════════════════════
# WORKER THREAD
# ═════════════════════════════════════════════════════════════════════════════

def run_job(job_id, cnae, provincia, n_paginas, ratio_min, ratio_max, delay):
    """
    Descarga N páginas del ranking nacional filtrado por CNAE.
    Divide las empresas en competidores (top del ranking) y leads (medianas).
    """
    def upd(status, pct, msg):
        jobs[job_id].update({"status": status, "progress": pct, "message": msg})

    try:
        session = make_session()
        upd("running", 5, "Conectando con el sitio…")

        # Warm-up para obtener cookies
        try:
            session.get("https://ranking-empresas.eleconomista.es/", timeout=15)
            time.sleep(1)
        except Exception:
            pass

        # Descargar todas las páginas solicitadas
        todas_empresas = []
        for p in range(1, n_paginas + 1):
            pct = 10 + int((p / n_paginas) * 65)
            upd("running", pct, f"Descargando página {p}/{n_paginas} del ranking…")
            soup, err = fetch_page(session, cnae, p)
            if err:
                upd("error", 0, err)
                return
            rows = parse_tabla(soup, provincia_filtro=provincia)
            if not rows and p == 1 and not provincia:
                upd("error", 0, f"No se encontraron empresas con CNAE {cnae}.")
                return
            todas_empresas.extend(rows)
            time.sleep(delay)

        if not todas_empresas:
            if provincia:
                upd("error", 0,
                    f"No hay empresas con CNAE {cnae} en {provincia} en las "
                    f"primeras {n_paginas} páginas. Aumenta el número de páginas.")
            else:
                upd("error", 0, "No se encontraron empresas para ese CNAE.")
            return

        upd("running", 80, f"Analizando {len(todas_empresas)} empresas…")

        # División Competidores / Leads
        # Competidores = top 25% por facturación (los más grandes del sector)
        # Leads        = el resto (medianas/pequeñas)
        # Si hay menos de 20 empresas, usar los primeros 5 como competidores
        empresas_con_fact = [e for e in todas_empresas if e["fact_num"]]
        empresas_con_fact.sort(key=lambda x: -x["fact_num"])

        n_total = len(empresas_con_fact)
        if n_total < 10:
            # Con pocos datos: primeros 3 = competidores, resto = leads
            competidores = empresas_con_fact[:3]
            leads = empresas_con_fact[3:] + [e for e in todas_empresas if not e["fact_num"]]
        else:
            # Top 25% = competidores, 75% = leads
            n_comp = max(5, n_total // 4)
            competidores = empresas_con_fact[:n_comp]
            leads_con_fact = empresas_con_fact[n_comp:]
            leads = leads_con_fact + [e for e in todas_empresas if not e["fact_num"]]

        # Para cada lead, buscar top 3 competidores
        upd("running", 90, "Emparejando leads con competidores…")
        resumen = []
        for lead in leads:
            comps = encontrar_competidores(lead, competidores, ratio_min, ratio_max, n=3)
            # Calcular ratio para cada uno
            for c in comps:
                if lead["fact_num"] and c["fact_num"]:
                    c["ratio"] = round(c["fact_num"] / lead["fact_num"], 1)
                else:
                    c["ratio"] = None
            resumen.append({"lead": lead, "competidores": comps})

        upd("running", 95, "Generando archivo Excel…")
        buf = build_excel(leads, resumen, cnae, provincia)

        jobs[job_id]["file"] = buf
        safe_prov = (provincia or "nacional").replace(" ", "_").lower()
        jobs[job_id]["filename"] = f"radar_cnae{cnae}_{safe_prov}.xlsx"

        upd("done", 100,
            f"✅ {len(leads)} leads · {len(competidores)} competidores")

    except Exception as e:
        import traceback
        upd("error", 0, f"Error inesperado: {type(e).__name__}: {e}")
        print(traceback.format_exc())


# ═════════════════════════════════════════════════════════════════════════════
# RUTAS FLASK
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html", provincias=PROVINCIAS)

@app.route("/start", methods=["POST"])
def start():
    data = request.json or {}
    cnae = str(data.get("cnae", "")).strip()
    if not cnae or not cnae.isdigit():
        return jsonify({"error": "CNAE obligatorio (solo dígitos)"}), 400

    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "pending", "progress": 0,
                    "message": "En cola…", "file": None}

    threading.Thread(
        target=run_job,
        daemon=True,
        args=(
            job_id,
            cnae,
            data.get("provincia", "").strip(),
            max(1, min(int(data.get("paginas", 3)), 10)),
            float(data.get("ratio_min", 3)),
            float(data.get("ratio_max", 20)),
            float(data.get("delay", 1.5)),
        ),
    ).start()
    return jsonify({"job_id": job_id})

@app.route("/status/<job_id>")
def status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "trabajo no encontrado"}), 404
    return jsonify({
        "status":   job["status"],
        "progress": job["progress"],
        "message":  job["message"],
        "ready":    job["status"] == "done",
    })

@app.route("/download/<job_id>")
def download(job_id):
    job = jobs.get(job_id)
    if not job or job["status"] != "done" or not job["file"]:
        return "Archivo no disponible", 404
    job["file"].seek(0)
    return send_file(
        job["file"],
        as_attachment=True,
        download_name=job.get("filename", "radar.xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/test")
def test():
    """Ruta de diagnóstico — abre http://localhost:5000/test"""
    import traceback
    out = {}
    try:
        session = make_session()
        session.get("https://ranking-empresas.eleconomista.es/", timeout=15)
        time.sleep(1)
        soup, err = fetch_page(session, "4662", 1)
        if err:
            out["error"] = err
        else:
            rows = parse_tabla(soup)
            out["status"] = "OK"
            out["empresas_encontradas"] = len(rows)
            out["primeras_5"] = [
                {"pos": r["posicion"], "nombre": r["nombre"],
                 "prov": r["provincia"], "fact": r["fact_label"]}
                for r in rows[:5]
            ]
    except Exception as e:
        out["error"] = str(e)
        out["trace"] = traceback.format_exc()
    return jsonify(out)


if __name__ == "__main__":
    print("\n" + "═" * 52)
    print("  RADAR COMPETIDOR  ·  http://localhost:5000")
    print("  Diagnóstico      ·  http://localhost:5000/test")
    print("═" * 52 + "\n")
    app.run(debug=False, port=5000, host="0.0.0.0")
