"""
Radar CRM — Aplicación Flask principal.
Ejecutar:  python app.py  →  http://localhost:5000
"""
import io
import threading
from datetime import datetime
from functools import wraps

from flask import (Flask, render_template, request, jsonify, redirect, url_for,
                   flash, send_file, abort)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from models import db, Usuario, Asignacion, Lead, Comentario, Actividad
from scraper import scrape_cnae
from enrichment import enrich_lead


# ── App ────────────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.config.from_object(config)

db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "Por favor inicia sesión para continuar."


@login_manager.user_loader
def load_user(uid):
    return db.session.get(Usuario, int(uid))


# ── Decoradores de rol ────────────────────────────────────────────────────────
def role_required(*roles):
    def deco(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("login"))
            if current_user.rol not in roles:
                abort(403)
            return f(*args, **kwargs)
        return wrapper
    return deco


# ═════════════════════════════════════════════════════════════════════════════
#  AUTH
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("home"))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        pwd   = request.form.get("password", "")
        user  = Usuario.query.filter_by(email=email).first()

        if user and user.activo and user.check_password(pwd):
            user.ultimo_acceso = datetime.utcnow()
            db.session.commit()
            login_user(user, remember=True)
            return redirect(url_for("home"))
        flash("Email o contraseña incorrectos.", "error")

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def home():
    if current_user.es_admin:      return redirect(url_for("admin_dashboard"))
    if current_user.es_supervisor: return redirect(url_for("supervisor_dashboard"))
    return redirect(url_for("comercial_kanban"))


# ═════════════════════════════════════════════════════════════════════════════
#  ADMIN
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/admin")
@role_required("admin")
def admin_dashboard():
    total_usuarios = Usuario.query.count()
    comerciales    = Usuario.query.filter_by(rol="comercial", activo=True).count()
    total_asigns   = Asignacion.query.count()
    total_leads    = Lead.query.count()

    asigns_recientes = (Asignacion.query
                        .order_by(Asignacion.fecha_creacion.desc())
                        .limit(10).all())

    # Distribución de leads por estado
    from sqlalchemy import func
    estados = (db.session.query(Lead.estado, func.count(Lead.id))
               .group_by(Lead.estado).all())

    return render_template("admin/dashboard.html",
                           total_usuarios=total_usuarios,
                           total_comerciales=comerciales,
                           total_asigns=total_asigns,
                           total_leads=total_leads,
                           asigns_recientes=asigns_recientes,
                           estados=dict(estados))


@app.route("/admin/usuarios")
@role_required("admin")
def admin_usuarios():
    usuarios = Usuario.query.order_by(Usuario.fecha_creacion.desc()).all()
    return render_template("admin/usuarios.html", usuarios=usuarios)


@app.route("/admin/usuarios/nuevo", methods=["POST"])
@role_required("admin")
def admin_usuario_nuevo():
    data = request.form
    email = data.get("email", "").strip().lower()
    if Usuario.query.filter_by(email=email).first():
        flash("Ya existe un usuario con ese email.", "error")
        return redirect(url_for("admin_usuarios"))

    u = Usuario(
        nombre = data.get("nombre", "").strip(),
        email  = email,
        rol    = data.get("rol", "comercial"),
        activo = True,
    )
    u.set_password(data.get("password", "cambiame"))
    db.session.add(u)
    db.session.commit()
    flash(f"Usuario «{u.nombre}» creado.", "success")
    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuarios/<int:uid>/toggle", methods=["POST"])
@role_required("admin")
def admin_usuario_toggle(uid):
    u = Usuario.query.get_or_404(uid)
    if u.id == current_user.id:
        flash("No puedes desactivarte a ti mismo.", "error")
    else:
        u.activo = not u.activo
        db.session.commit()
        flash(f"Usuario {'activado' if u.activo else 'desactivado'}.", "success")
    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuarios/<int:uid>/eliminar", methods=["POST"])
@role_required("admin")
def admin_usuario_eliminar(uid):
    u = Usuario.query.get_or_404(uid)
    if u.id == current_user.id:
        flash("No puedes eliminarte a ti mismo.", "error")
    else:
        db.session.delete(u)
        db.session.commit()
        flash("Usuario eliminado.", "success")
    return redirect(url_for("admin_usuarios"))


@app.route("/admin/asignaciones")
@role_required("admin")
def admin_asignaciones():
    asigns = (Asignacion.query
              .order_by(Asignacion.fecha_creacion.desc()).all())
    comerciales = Usuario.query.filter_by(rol="comercial", activo=True).all()
    return render_template("admin/asignaciones.html",
                           asigns=asigns, comerciales=comerciales)


@app.route("/admin/asignaciones/nueva", methods=["POST"])
@role_required("admin")
def admin_asignacion_nueva():
    data = request.form
    try:
        comercial_id = int(data["comercial_id"])
    except (KeyError, ValueError):
        flash("Selecciona un comercial.", "error")
        return redirect(url_for("admin_asignaciones"))

    comercial = Usuario.query.get_or_404(comercial_id)
    cnae = data.get("cnae", "").strip()
    if not cnae.isdigit() or len(cnae) < 3:
        flash("CNAE inválido.", "error")
        return redirect(url_for("admin_asignaciones"))

    asig = Asignacion(
        comercial_id  = comercial.id,
        creado_por_id = current_user.id,
        cnae          = cnae,
        cnae_desc     = data.get("cnae_desc", "").strip(),
        provincia     = data.get("provincia", "").strip(),
        paginas       = max(1, min(int(data.get("paginas", 3) or 3),
                                    config.SCRAPE_MAX_PAGES)),
    )
    db.session.add(asig)
    db.session.commit()

    # Lanzar worker en background
    threading.Thread(target=_run_scrape_job,
                     args=(app, asig.id),
                     daemon=True).start()

    flash(f"Asignación creada. Iniciando scraping…", "success")
    return redirect(url_for("admin_asignaciones"))


@app.route("/admin/asignaciones/<int:aid>/estado")
@login_required
def asignacion_estado(aid):
    a = Asignacion.query.get_or_404(aid)
    return jsonify({
        "estado":    a.estado,
        "progreso":  a.progreso,
        "mensaje":   a.mensaje,
        "total":     a.total_leads,
    })


@app.route("/admin/asignaciones/<int:aid>/eliminar", methods=["POST"])
@role_required("admin")
def admin_asignacion_eliminar(aid):
    a = Asignacion.query.get_or_404(aid)
    db.session.delete(a)
    db.session.commit()
    flash("Asignación eliminada (con todos sus leads).", "success")
    return redirect(url_for("admin_asignaciones"))


@app.route("/admin/export/leads.xlsx")
@role_required("admin")
def admin_export_leads():
    """Exporta TODOS los leads a Excel."""
    leads = Lead.query.all()
    buf = _build_excel_leads(leads)
    return send_file(buf, as_attachment=True,
                     download_name=f"radar_leads_{datetime.now():%Y%m%d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ═════════════════════════════════════════════════════════════════════════════
#  SUPERVISOR
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/supervisor")
@role_required("admin", "supervisor")
def supervisor_dashboard():
    from sqlalchemy import func
    comerciales = Usuario.query.filter_by(rol="comercial", activo=True).all()

    # Stats por comercial
    stats = []
    for c in comerciales:
        total = Lead.query.filter_by(comercial_id=c.id).count()
        if total == 0: continue
        estados_counts = dict(
            db.session.query(Lead.estado, func.count(Lead.id))
            .filter(Lead.comercial_id == c.id)
            .group_by(Lead.estado).all()
        )
        act_recientes = (Actividad.query
                         .filter_by(usuario_id=c.id)
                         .order_by(Actividad.fecha.desc())
                         .limit(5).all())
        stats.append({
            "comercial": c,
            "total":     total,
            "estados":   estados_counts,
            "actividad": act_recientes,
            "ganados":   estados_counts.get("ganado", 0),
            "perdidos":  estados_counts.get("perdido", 0),
        })

    # Actividad reciente global
    actividad_global = (Actividad.query
                        .order_by(Actividad.fecha.desc())
                        .limit(30).all())

    return render_template("supervisor/dashboard.html",
                           stats=stats,
                           actividad_global=actividad_global)


# ═════════════════════════════════════════════════════════════════════════════
#  COMERCIAL (Kanban)
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/kanban")
@login_required
def comercial_kanban():
    # Admin y supervisor pueden ver kanban de un comercial con ?user_id=X
    user_id = request.args.get("user_id", type=int)
    if user_id and current_user.rol in ("admin", "supervisor"):
        target = Usuario.query.get_or_404(user_id)
    else:
        target = current_user

    leads = (Lead.query.filter_by(comercial_id=target.id)
             .order_by(Lead.orden, Lead.id).all())

    # Agrupar por estado
    por_estado = {k: [] for k, _, _ in config.KANBAN_ESTADOS}
    for l in leads:
        por_estado.setdefault(l.estado, []).append(l)

    return render_template("comercial/kanban.html",
                           target=target,
                           por_estado=por_estado,
                           estados=config.KANBAN_ESTADOS,
                           total=len(leads))


@app.route("/lead/<int:lid>")
@login_required
def lead_detail(lid):
    lead = Lead.query.get_or_404(lid)
    if not (current_user.rol in ("admin", "supervisor")
            or lead.comercial_id == current_user.id):
        abort(403)
    return render_template("comercial/lead_detail.html",
                           lead=lead,
                           estados=config.KANBAN_ESTADOS)


@app.route("/lead/<int:lid>/estado", methods=["POST"])
@login_required
def lead_cambiar_estado(lid):
    lead = Lead.query.get_or_404(lid)
    if not (current_user.rol == "admin" or lead.comercial_id == current_user.id):
        abort(403)

    nuevo = request.json.get("estado")
    if nuevo not in [k for k, _, _ in config.KANBAN_ESTADOS]:
        return jsonify({"error": "estado inválido"}), 400

    anterior = lead.estado
    lead.estado = nuevo
    db.session.add(Actividad(
        lead_id    = lead.id,
        usuario_id = current_user.id,
        tipo       = "estado_cambio",
        detalle    = f"{anterior} → {nuevo}",
    ))
    db.session.commit()
    return jsonify({"ok": True})


@app.route("/lead/<int:lid>/comentario", methods=["POST"])
@login_required
def lead_comentario(lid):
    lead = Lead.query.get_or_404(lid)
    if not (current_user.rol == "admin" or lead.comercial_id == current_user.id):
        abort(403)

    texto = (request.form.get("texto") or "").strip()
    if not texto:
        return jsonify({"error": "vacío"}), 400

    c = Comentario(lead_id=lead.id, autor_id=current_user.id, texto=texto)
    db.session.add(c)
    db.session.add(Actividad(
        lead_id    = lead.id,
        usuario_id = current_user.id,
        tipo       = "comentario",
        detalle    = texto[:100],
    ))
    db.session.commit()
    return jsonify({
        "id":    c.id,
        "texto": c.texto,
        "autor": current_user.nombre,
        "fecha": c.fecha.strftime("%d/%m/%Y %H:%M"),
    })


@app.route("/lead/<int:lid>/editar", methods=["POST"])
@login_required
def lead_editar(lid):
    """Permite editar campos extra (teléfono, email, web...) manualmente."""
    lead = Lead.query.get_or_404(lid)
    if not (current_user.rol == "admin" or lead.comercial_id == current_user.id):
        abort(403)

    data = request.json or {}
    campos_editables = ("telefono", "email", "web", "direccion", "gerente", "licita")
    for k in campos_editables:
        if k in data:
            setattr(lead, k, data[k].strip() if data[k] else None)

    db.session.add(Actividad(
        lead_id    = lead.id,
        usuario_id = current_user.id,
        tipo       = "edicion",
        detalle    = f"Datos actualizados: {', '.join(data.keys())}",
    ))
    db.session.commit()
    return jsonify({"ok": True})


@app.route("/lead/<int:lid>/enriquecer", methods=["POST"])
@login_required
def lead_enriquecer(lid):
    """Dispara re-enriquecimiento manual de un lead."""
    lead = Lead.query.get_or_404(lid)
    if not (current_user.rol in ("admin", "supervisor")
            or lead.comercial_id == current_user.id):
        abort(403)

    threading.Thread(target=_enrich_single_lead,
                     args=(app, lead.id),
                     daemon=True).start()
    return jsonify({"ok": True, "mensaje": "Enriquecimiento lanzado"})


# ═════════════════════════════════════════════════════════════════════════════
#  WORKERS (background threads)
# ═════════════════════════════════════════════════════════════════════════════

def _run_scrape_job(flask_app, asignacion_id):
    """Descarga leads de la asignación + lanza enriquecimiento."""
    with flask_app.app_context():
        asig = db.session.get(Asignacion, asignacion_id)
        if not asig:
            return

        def set_estado(estado, progreso=None, msg=None):
            asig.estado   = estado
            if progreso is not None: asig.progreso = progreso
            if msg:                  asig.mensaje  = msg
            db.session.commit()

        set_estado("scrapeando", 5, "Conectando con el ranking…")

        def on_progress(pct, msg):
            # 5-60% = scraping inicial
            asig.progreso = 5 + int(pct * 0.55)
            asig.mensaje  = msg
            db.session.commit()

        empresas, err = scrape_cnae(
            cnae      = asig.cnae,
            provincia = asig.provincia or None,
            paginas   = asig.paginas,
            delay     = config.SCRAPE_DELAY_SECONDS,
            on_progress = on_progress,
        )

        if err:
            set_estado("error", 0, f"Error scraping: {err}")
            return

        if not empresas:
            if asig.provincia:
                set_estado("error", 0,
                    f"No se encontraron empresas con CNAE {asig.cnae} en "
                    f"{asig.provincia} en {asig.paginas} página(s). "
                    f"Prueba aumentando el nº de páginas — las provincias pequeñas "
                    f"aparecen más adelante en el ranking nacional.")
            else:
                set_estado("error", 0,
                    f"No se encontraron empresas con CNAE {asig.cnae}. "
                    f"Verifica que el código CNAE sea correcto.")
            return

        # Crear leads en BD
        set_estado("scrapeando", 60, f"Guardando {len(empresas)} leads…")
        for e in empresas:
            lead = Lead(
                asignacion_id     = asig.id,
                comercial_id      = asig.comercial_id,
                nombre            = e["nombre"],
                cnae              = e["cnae"],
                provincia         = e["provincia"],
                posicion_nacional = e["posicion"],
                evolucion         = e["evolucion"],
                tendencia         = e["tendencia"],
                facturacion_num   = e["facturacion_num"],
                facturacion_raw   = e["facturacion_raw"],
                url_ficha         = e["url"],
                estado            = "nuevo",
            )
            db.session.add(lead)
        asig.total_leads = len(empresas)
        db.session.commit()

        # Enriquecimiento en background (no bloquear)
        set_estado("completada", 100,
                   f"✅ {len(empresas)} leads creados. Enriqueciendo en background…")
        asig.fecha_completada = datetime.utcnow()
        db.session.commit()

        # Lanzar enriquecimiento por lotes pequeños
        lead_ids = [l.id for l in asig.leads.all()]
        threading.Thread(target=_enrich_batch,
                         args=(flask_app, lead_ids),
                         daemon=True).start()


def _enrich_single_lead(flask_app, lead_id):
    with flask_app.app_context():
        lead = db.session.get(Lead, lead_id)
        if not lead: return

        datos = enrich_lead({
            "nombre":    lead.nombre,
            "provincia": lead.provincia,
            "url":       lead.url_ficha,
        })
        for k, v in datos.items():
            if v and not getattr(lead, k, None):
                setattr(lead, k, v)
        lead.enriquecido = True
        db.session.commit()


def _enrich_batch(flask_app, lead_ids):
    """Enriquece leads uno a uno con pausas (para no saturar Google)."""
    for i, lid in enumerate(lead_ids):
        try:
            _enrich_single_lead(flask_app, lid)
        except Exception as e:
            print(f"[enrich_batch] lead {lid}: {e}")
        # Pausa entre peticiones
        import time; time.sleep(2)


# ═════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ═════════════════════════════════════════════════════════════════════════════

def _build_excel_leads(leads):
    C_HDR = "8B1A1A"; C_TITLE = "5C0F0F"; C_EVEN = "FDECEC"; C_WHITE = "FFFFFF"
    THIN = Border(*[Side(style="thin", color="E5D3D3")]*4)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.sheet_properties.tabColor = C_HDR

    # Título
    ws.merge_cells("A1:N1")
    c = ws.cell(1, 1, f"RADAR CRM · Export de Leads · {datetime.now():%d/%m/%Y}")
    c.font = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=C_TITLE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    headers = ["Empresa","CNAE","Provincia","Pos.","Facturación","Tendencia",
               "Teléfono","Email","Web","Dirección","Gerente","Licita",
               "Estado","Comercial"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(3, i, h)
        cell.font = Font(name="Calibri", bold=True, color=C_WHITE)
        cell.fill = PatternFill("solid", start_color=C_HDR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
    ws.row_dimensions[3].height = 26

    for r, l in enumerate(leads, 4):
        bg = C_EVEN if r % 2 == 0 else C_WHITE
        vals = [l.nombre, l.cnae, l.provincia, l.posicion_nacional,
                l.facturacion_raw, l.tendencia, l.telefono, l.email, l.web,
                l.direccion, l.gerente, l.licita, l.estado,
                l.comercial.nombre if l.comercial else ""]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(r, i, v or "")
            cell.font = Font(name="Calibri", size=10)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="left" if i in (1,7,8,9,10,11) else "center")
            cell.border = THIN

    widths = [42,7,14,8,18,10,14,28,28,32,26,9,14,20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:N{3+len(leads)}"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ═════════════════════════════════════════════════════════════════════════════
#  CONTEXT PROCESSORS / FILTERS
# ═════════════════════════════════════════════════════════════════════════════

@app.context_processor
def inject_globals():
    return {
        "KANBAN_ESTADOS": config.KANBAN_ESTADOS,
        "ROLES":          config.ROLES,
    }


@app.template_filter("fmt_fecha")
def fmt_fecha(f):
    return f.strftime("%d/%m/%Y %H:%M") if f else ""


@app.template_filter("fmt_num")
def fmt_num(n):
    if n is None: return "—"
    try: return f"{int(n):,}".replace(",", ".")
    except: return str(n)


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("\n" + "="*54)
    print("  RADAR CRM  ·  http://localhost:5000")
    print("="*54 + "\n")
    app.run(debug=False, port=5000, host="0.0.0.0")
