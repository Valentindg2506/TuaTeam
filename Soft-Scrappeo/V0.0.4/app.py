"""
Radar CRM v2 — Flask app completa.
Novedades: competidores en lead, notificaciones, chat, bugs corregidos.
"""
import io, threading, traceback, time, re
import datetime as _dt
from datetime import datetime, timezone

def utcnow():
    """Replacement for deprecated datetime.utcnow()."""
    return datetime.now(timezone.utc).replace(tzinfo=None)
from functools import wraps
import os
from dotenv import load_dotenv
load_dotenv()

from flask import (Flask, render_template, request, jsonify, redirect,
                   url_for, flash, send_file, abort)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from models import (db, Usuario, Asignacion, Lead, Competidor,
                    Comentario, Actividad, Notificacion, MensajeChat)
from scraper import scrape_cnae, calcular_competidores
from enrichment import enrich_lead
from cnae_catalog import CNAE_CATALOG

app = Flask(__name__)
app.config.from_object(config)
db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "Inicia sesión para continuar."

# Control global anti-bloqueo para scraping concurrente.
SCRAPE_MAX_CONCURRENT = max(1, int(getattr(config, "SCRAPE_MAX_CONCURRENT_JOBS", 2)))
_SCRAPE_SLOT = threading.Semaphore(SCRAPE_MAX_CONCURRENT)

@login_manager.user_loader
def load_user(uid): return db.session.get(Usuario, int(uid))

# ── Decoradores ───────────────────────────────────────────────────────────────
def role_required(*roles):
    def deco(f):
        @wraps(f)
        def w(*a, **kw):
            if not current_user.is_authenticated: return redirect(url_for("login"))
            if current_user.rol not in roles: abort(403)
            return f(*a, **kw)
        return w
    return deco

# ── Helpers ───────────────────────────────────────────────────────────────────
def crear_notif(usuario_id, tipo, titulo, texto="", url=""):
    n = Notificacion(usuario_id=usuario_id, tipo=tipo,
                     titulo=titulo, texto=texto, url=url)
    db.session.add(n)

def _norm_phone(v):
    s = re.sub(r"\D", "", str(v or ""))
    if s.startswith("0034"): s = s[4:]
    if s.startswith("34") and len(s) > 9: s = s[2:]
    return s[:9] if len(s) >= 9 else (s or None)

def _norm_email(v):
    e = (v or "").strip().lower()
    return e or None

def _norm_web(v):
    w = (v or "").strip()
    if not w:
        return None
    if w.startswith("//"):
        w = "https:" + w
    elif not w.startswith(("http://", "https://")):
        w = "https://" + w
    return w

def _norm_gerente(v):
    g = re.sub(r"\s+", " ", str(v or "")).strip(" .,:;|-")
    if not g:
        return None
    low = g.lower()
    if any(x in low for x in (
        "sitio web", "pagina web", "página web", "website",
        "aviso legal", "politica de privacidad", "política de privacidad",
        "cookies", "contacto",
    )):
        return None

    toks = re.findall(r"[A-Za-zÁÉÍÓÚÑáéíóúñ]+", g)
    if len(toks) < 2 or len(toks) > 6:
        return None
    return g[:180]

def _lead_completeness_map(data):
    k = ("telefono", "email", "web", "direccion", "gerente")
    return sum(1 for x in k if (data.get(x) or "").strip())

_CNAE_CACHE = {"ts": 0.0, "data": {}}
_CNAE_CACHE_LOCK = threading.Lock()

def _get_cnae_catalog():
    """
    Catalogo para autocompletado CNAE:
    - Base estatica (curada)
    - Refuerzo dinamico desde asignaciones con descripcion
    Cacheado para evitar consultas de DB por cada tecla.
    """
    now = time.time()
    with _CNAE_CACHE_LOCK:
        if _CNAE_CACHE["data"] and (now - _CNAE_CACHE["ts"] < 600):
            return _CNAE_CACHE["data"]

    data = dict(CNAE_CATALOG)
    try:
        rows = (db.session.query(Asignacion.cnae, Asignacion.cnae_desc)
                .filter(Asignacion.cnae != None, Asignacion.cnae_desc != None)
                .all())
        for code, desc in rows:
            c = re.sub(r"\D", "", str(code or "")).strip()
            d = (desc or "").strip()
            if 3 <= len(c) <= 5 and d:
                data[c] = d
    except Exception:
        pass

    with _CNAE_CACHE_LOCK:
        _CNAE_CACHE["ts"] = now
        _CNAE_CACHE["data"] = data
    return data

# ═════════════════════════════════════════════════════════════════════════════
#  AUTH
# ═════════════════════════════════════════════════════════════════════════════
@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated: return redirect(url_for("home"))
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        pwd   = request.form.get("password", "")
        user  = Usuario.query.filter_by(email=email).first()
        if user and user.activo and user.check_password(pwd):
            user.ultimo_acceso = utcnow()
            db.session.commit()
            login_user(user, remember=True)
            return redirect(url_for("home"))
        flash("Email o contraseña incorrectos.", "error")
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def home():
    if current_user.es_admin:      return redirect(url_for("admin_dashboard"))
    if current_user.es_supervisor: return redirect(url_for("supervisor_dashboard"))
    return redirect(url_for("comercial_kanban"))

# ═════════════════════════════════════════════════════════════════════════════
#  ADMIN
# ═════════════════════════════════════════════════════════════════════════════
@app.route("/admin")
@role_required("admin")
def admin_dashboard():
    from sqlalchemy import func
    # Variable names match dashboard.html template expectations
    total_leads      = Lead.query.count()
    total_asigns     = Asignacion.query.count()
    total_comerciales= Usuario.query.filter_by(rol="comercial", activo=True).count()
    total_usuarios   = Usuario.query.count()
    estados = dict(db.session.query(Lead.estado, func.count(Lead.id))
                   .group_by(Lead.estado).all())
    asigns_recientes = Asignacion.query.order_by(Asignacion.fecha_creacion.desc()).limit(10).all()
    return render_template("admin/dashboard.html",
                           total_leads=total_leads,
                           total_asigns=total_asigns,
                           total_comerciales=total_comerciales,
                           total_usuarios=total_usuarios,
                           estados=estados,
                           asigns_recientes=asigns_recientes)

@app.route("/admin/usuarios")
@role_required("admin")
def admin_usuarios():
    return render_template("admin/usuarios.html",
                           usuarios=Usuario.query.order_by(Usuario.fecha_creacion.desc()).all())

@app.route("/admin/usuarios/nuevo", methods=["POST"])
@role_required("admin")
def admin_usuario_nuevo():
    email = request.form.get("email","").strip().lower()
    if Usuario.query.filter_by(email=email).first():
        flash("Ya existe ese email.", "error")
        return redirect(url_for("admin_usuarios"))
    u = Usuario(nombre=request.form.get("nombre","").strip(),
                email=email, rol=request.form.get("rol","comercial"), activo=True)
    u.set_password(request.form.get("password","cambiame"))
    db.session.add(u)
    db.session.commit()
    flash(f"Usuario «{u.nombre}» creado.", "success")
    return redirect(url_for("admin_usuarios"))

@app.route("/admin/usuarios/<int:uid>/toggle", methods=["POST"])
@role_required("admin")
def admin_usuario_toggle(uid):
    u = db.session.get(Usuario, uid) or abort(404)
    if u.id != current_user.id:
        u.activo = not u.activo
        db.session.commit()
        flash(f"Usuario {'activado' if u.activo else 'desactivado'}.", "success")
    return redirect(url_for("admin_usuarios"))

@app.route("/admin/usuarios/<int:uid>/eliminar", methods=["POST"])
@role_required("admin")
def admin_usuario_eliminar(uid):
    u = db.session.get(Usuario, uid) or abort(404)
    if u.id != current_user.id:
        db.session.delete(u); db.session.commit()
        flash("Usuario eliminado.", "success")
    return redirect(url_for("admin_usuarios"))

@app.route("/admin/asignaciones")
@role_required("admin")
def admin_asignaciones():
    asigns = Asignacion.query.order_by(Asignacion.fecha_creacion.desc()).all()
    comerciales = Usuario.query.filter_by(rol="comercial", activo=True).all()
    return render_template("admin/asignaciones.html",
                           asigns=asigns, comerciales=comerciales)

@app.route("/admin/asignaciones/nueva", methods=["POST"])
@role_required("admin")
def admin_asignacion_nueva():
    try: cid = int(request.form["comercial_id"])
    except: flash("Selecciona un comercial.", "error"); return redirect(url_for("admin_asignaciones"))

    cnae = request.form.get("cnae","").strip()
    if not cnae.isdigit() or len(cnae) < 3:
        flash("CNAE inválido.", "error"); return redirect(url_for("admin_asignaciones"))

    cnae_desc = (request.form.get("cnae_desc","") or "").strip()
    if not cnae_desc:
        cnae_desc = (_get_cnae_catalog().get(cnae) or "").strip()

    asig = Asignacion(
        comercial_id=cid, creado_por_id=current_user.id,
        cnae=cnae, cnae_desc=cnae_desc,
        provincia=request.form.get("provincia","").strip(),
        paginas=0,  # 0 = modo exhaustivo automático
    )
    db.session.add(asig); db.session.commit()
    threading.Thread(target=_run_scrape, args=(app, asig.id), daemon=True).start()
    flash("Asignación creada. Scraping exhaustivo iniciado en segundo plano.", "success")
    return redirect(url_for("admin_asignaciones"))

@app.route("/admin/asignaciones/<int:aid>/estado")
@login_required
def asignacion_estado_api(aid):
    a = db.session.get(Asignacion, aid) or abort(404)
    return jsonify({"estado": a.estado, "progreso": a.progreso,
                    "mensaje": a.mensaje, "total": a.total_leads})

@app.route("/api/cnae/sugerencias")
@role_required("admin")
def api_cnae_sugerencias():
    q_raw = (request.args.get("q") or "").strip()
    q_digits = re.sub(r"\D", "", q_raw)
    q_text = q_raw.lower()
    if not q_digits and len(q_text) < 2:
        return jsonify([])

    catalog = _get_cnae_catalog()
    items = []
    for code, desc in catalog.items():
        c = str(code or "").strip()
        d = str(desc or "").strip()
        if not c or not d:
            continue
        if q_digits and c.startswith(q_digits):
            rank = 0
        elif q_text and q_text in d.lower():
            rank = 1
        elif q_digits and q_digits in c:
            rank = 2
        else:
            continue
        items.append((rank, c, d))

    items.sort(key=lambda x: (x[0], len(x[1]), x[1]))
    out = [{
        "cnae": c,
        "descripcion": d,
        "label": f"{c} - {d}",
    } for _, c, d in items[:15]]
    return jsonify(out)

@app.route("/api/dashboard/stats")
@role_required("admin", "supervisor")
def api_dashboard_stats():
    """Stats en tiempo real para el dashboard."""
    from sqlalchemy import func
    return jsonify({
        "total_leads":       Lead.query.count(),
        "total_asigns":      Asignacion.query.count(),
        "total_comerciales": Usuario.query.filter_by(rol="comercial", activo=True).count(),
        "total_usuarios":    Usuario.query.count(),
        "estados": dict(db.session.query(Lead.estado, func.count(Lead.id))
                        .group_by(Lead.estado).all()),
    })

@app.route("/api/asignaciones/activas")
@login_required
def api_asignaciones_activas():
    """Retorna IDs + estado de asignaciones en curso o solicitadas (para polling)."""
    # Si el cliente pide IDs específicos, se devuelven todos esos sin filtrar por estado
    ids_param = request.args.get("ids")
    if ids_param:
        ids_list = [int(x) for x in ids_param.split(",") if x.isdigit()]
        activas = Asignacion.query.filter(Asignacion.id.in_(ids_list)).all()
    else:
        activas = Asignacion.query.filter(
            Asignacion.estado.in_(["scrapeando", "pendiente"])
        ).all()
        
    return jsonify([{
        "id": a.id, "estado": a.estado,
        "progreso": a.progreso, "mensaje": a.mensaje, "total": a.total_leads
    } for a in activas])


@app.route("/admin/asignaciones/<int:aid>/eliminar", methods=["POST"])
@role_required("admin")
def admin_asignacion_eliminar(aid):
    a = db.session.get(Asignacion, aid) or abort(404)
    db.session.delete(a); db.session.commit()
    flash("Asignación eliminada.", "success")
    return redirect(url_for("admin_asignaciones"))

@app.route("/admin/asignaciones/<int:aid>/reintentar", methods=["POST"])
@role_required("admin")
def admin_asignacion_reintentar(aid):
    a = db.session.get(Asignacion, aid) or abort(404)
    if a.estado == "scrapeando":
        flash("La asignación ya está scrapeando.", "error")
        return redirect(url_for("admin_asignaciones"))

    # Reiniciar estado y limpiar resultados previos de esta asignación.
    for lead in a.leads.all():
        db.session.delete(lead)

    a.estado = "pendiente"
    a.progreso = 0
    a.mensaje = "Reintento manual solicitado…"
    a.total_leads = 0
    a.paginas = 0  # Forzar modo exhaustivo en reintentos
    a.fecha_completada = None
    db.session.commit()

    threading.Thread(target=_run_scrape, args=(app, a.id), daemon=True).start()
    flash("Reintento de scraping iniciado.", "success")
    return redirect(url_for("admin_asignaciones"))

@app.route("/admin/export/leads.xlsx")
@role_required("admin")
def admin_export_leads():
    leads = Lead.query.all()
    return send_file(_build_excel(leads), as_attachment=True,
                     download_name=f"radar_{datetime.now():%Y%m%d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

g_reenrich = {"activo": False, "total": 0, "hechos": 0}

@app.route("/admin/re-enriquecer-faltantes", methods=["POST"])
@role_required("admin")
def admin_re_enriquecer_faltantes():
    global g_reenrich
    from sqlalchemy import or_
    if g_reenrich["activo"]:
        return jsonify({"ok": False, "msg": "Ya hay un proceso de re-enriquecimiento en curso."})
        
    from enrichment import enrich_lead
    leads = Lead.query.filter(
        or_(Lead.telefono == None, Lead.email == None, Lead.telefono == "", Lead.email == "", Lead.web == None)
    ).all()
    
    if not leads:
        return jsonify({"ok": True, "msg": "Todos los leads ya tienen datos. No hay nada que re-enriquecer."})
        
    for l in leads:
        l.enriquecido = False
    db.session.commit()

    g_reenrich["activo"] = True
    g_reenrich["total"] = len(leads)
    g_reenrich["hechos"] = 0

    def _batch_enrich_missing(flask_app, lids):
        global g_reenrich
        with flask_app.app_context():
            import time
            for lid in lids:
                l = db.session.get(Lead, lid)
                if l:
                    try: 
                        _enrich_one(flask_app, lid)
                    except Exception: 
                        pass
                g_reenrich["hechos"] += 1
                time.sleep(2)
        g_reenrich["activo"] = False
                
    threading.Thread(target=_batch_enrich_missing, args=(app, [l.id for l in leads]), daemon=True).start()
    
    return jsonify({"ok": True, "msg": f"Se ha iniciado el escaneo profundo de {len(leads)} leads incompletos en segundo plano."})

@app.route("/admin/re-enriquecer-estado")
@role_required("admin")
def admin_reenriquecer_estado():
    if not g_reenrich["activo"]:
        return jsonify({"activo": False})
    
    pct = int((g_reenrich["hechos"] / g_reenrich["total"]) * 100) if g_reenrich["total"] > 0 else 0
    return jsonify({
        "activo": True,
        "total": g_reenrich["total"],
        "hechos": g_reenrich["hechos"],
        "pct": pct
    })

# ═════════════════════════════════════════════════════════════════════════════
#  SUPERVISOR
# ═════════════════════════════════════════════════════════════════════════════
@app.route("/supervisor")
@role_required("admin","supervisor")
def supervisor_dashboard():
    from sqlalchemy import func
    comerciales = Usuario.query.filter_by(rol="comercial", activo=True).all()
    stats = []
    for c in comerciales:
        total = Lead.query.filter_by(comercial_id=c.id).count()
        if total == 0: continue
        estados = dict(db.session.query(Lead.estado, func.count(Lead.id))
                       .filter(Lead.comercial_id==c.id).group_by(Lead.estado).all())
        acts = (Actividad.query.filter_by(usuario_id=c.id)
                .order_by(Actividad.fecha.desc()).limit(5).all())
        stats.append({"comercial":c,"total":total,"estados":estados,"actividad":acts,
                      "ganados":estados.get("ganado",0),"perdidos":estados.get("perdido",0)})

    actividad = Actividad.query.order_by(Actividad.fecha.desc()).limit(30).all()
    return render_template(
        "supervisor/dashboard.html",
        stats=stats,
        actividad=actividad,
        actividad_global=actividad,
    )

# ═════════════════════════════════════════════════════════════════════════════
#  KANBAN
# ═════════════════════════════════════════════════════════════════════════════
@app.route("/kanban")
@login_required
def comercial_kanban():
    from sqlalchemy import or_

    uid = request.args.get("user_id", type=int)
    target = (db.session.get(Usuario, uid)
              if uid and current_user.rol in ("admin","supervisor")
              else current_user)
    if not target: abort(404)

    vista = (request.args.get("vista") or "kanban").strip().lower()
    if vista not in ("kanban", "lista"):
        vista = "kanban"

    q = (request.args.get("q") or "").strip()
    leads_q = Lead.query.filter_by(comercial_id=target.id)
    if q:
        like_q = f"%{q}%"
        leads_q = leads_q.filter(or_(
            Lead.nombre.ilike(like_q),
            Lead.cnae.ilike(like_q),
            Lead.provincia.ilike(like_q),
            Lead.email.ilike(like_q),
            Lead.telefono.ilike(like_q),
        ))
    leads = leads_q.order_by(Lead.orden, Lead.id).all()

    por_estado = {k:[] for k,_,_ in config.KANBAN_ESTADOS}
    for l in leads:
        por_estado.setdefault(l.estado, []).append(l)

    return render_template(
        "comercial/kanban.html",
        target=target,
        leads=leads,
        por_estado=por_estado,
        estados=config.KANBAN_ESTADOS,
        total=len(leads),
        vista=vista,
        q=q,
    )

# ═════════════════════════════════════════════════════════════════════════════
#  LEAD DETAIL
# ═════════════════════════════════════════════════════════════════════════════
@app.route("/lead/<int:lid>")
@login_required
def lead_detail(lid):
    lead = Lead.query.get_or_404(lid)
    if not current_user.puede_ver_lead(lead): abort(403)
    return render_template("comercial/lead_detail.html",
                           lead=lead, estados=config.KANBAN_ESTADOS)

@app.route("/lead/<int:lid>/estado", methods=["POST"])
@login_required
def lead_cambiar_estado(lid):
    lead = Lead.query.get_or_404(lid)
    if not current_user.puede_ver_lead(lead): abort(403)
    nuevo = (request.json or {}).get("estado")
    if nuevo not in [k for k,_,_ in config.KANBAN_ESTADOS]:
        return jsonify({"error":"estado inválido"}), 400
    anterior = lead.estado
    lead.estado = nuevo
    lead.fecha_actualizacion = utcnow()
    db.session.add(Actividad(lead_id=lead.id, usuario_id=current_user.id,
                             tipo="estado_cambio", detalle=f"{anterior} → {nuevo}"))
    # Notificar al supervisor si el lead se gana o pierde
    if nuevo in ("ganado","perdido"):
        svs = Usuario.query.filter_by(rol="supervisor", activo=True).all()
        admins = Usuario.query.filter_by(rol="admin", activo=True).all()
        emoji = "🏆" if nuevo == "ganado" else "❌"
        for u in svs + admins:
            crear_notif(u.id, "lead",
                f"{emoji} Lead {nuevo}: {lead.nombre[:30]}",
                f"Por {current_user.nombre}",
                url_for("lead_detail", lid=lead.id))
    db.session.commit()
    return jsonify({"ok":True, "estado":nuevo})

@app.route("/lead/<int:lid>/comentario", methods=["POST"])
@login_required
def lead_comentario(lid):
    lead = Lead.query.get_or_404(lid)
    if not current_user.puede_ver_lead(lead): abort(403)
    texto = (request.form.get("texto") or "").strip()
    if not texto: return jsonify({"error":"vacío"}), 400
    c = Comentario(lead_id=lid, autor_id=current_user.id, texto=texto)
    db.session.add(c)
    db.session.add(Actividad(lead_id=lid, usuario_id=current_user.id,
                             tipo="comentario", detalle=texto[:100]))
    db.session.commit()
    return jsonify({"id":c.id,"texto":c.texto,"autor":current_user.nombre,
                    "fecha":c.fecha.strftime("%d/%m/%Y %H:%M")})

@app.route("/lead/<int:lid>/editar", methods=["POST"])
@login_required
def lead_editar(lid):
    lead = Lead.query.get_or_404(lid)
    if not current_user.puede_ver_lead(lead): abort(403)
    data = request.json or {}
    for k in ("telefono","email","web","direccion","gerente","licita"):
        if k in data:
            setattr(lead, k, (data[k] or "").strip() or None)
    lead.fecha_actualizacion = utcnow()
    db.session.add(Actividad(lead_id=lid, usuario_id=current_user.id,
                             tipo="edicion", detalle=f"Editado: {', '.join(data)}"))
    db.session.commit()
    return jsonify({"ok":True})

@app.route("/lead/<int:lid>/enriquecer", methods=["POST"])
@login_required
def lead_enriquecer(lid):
    lead = Lead.query.get_or_404(lid)
    if not current_user.puede_ver_lead(lead): abort(403)
    
    # Reiniciar la bandera para que el frontend sepa que debe esperar
    lead.enriquecido = False
    db.session.commit()
    
    threading.Thread(target=_enrich_one, args=(app, lid), daemon=True).start()
    return jsonify({"ok": True, "msg": "Enriquecimiento iniciado. Los datos aparecerán en ~30s."})

@app.route("/api/lead/<int:lid>/datos")
@login_required
def api_lead_datos(lid):
    """Devuelve los datos de contacto del lead para actualización sin recarga."""
    lead = Lead.query.get_or_404(lid)
    if not current_user.puede_ver_lead(lead): abort(403)
    return jsonify({
        "telefono":  lead.telefono,
        "email":     lead.email,
        "web":       lead.web,
        "direccion": lead.direccion,
        "gerente":   lead.gerente,
        "licita":    lead.licita,
        "enriquecido": lead.enriquecido,
    })

@app.route("/lead/<int:lid>/comentario/<int:cid>/eliminar", methods=["POST"])
@login_required
def lead_comentario_eliminar(lid, cid):
    c = Comentario.query.get_or_404(cid)
    if c.autor_id != current_user.id and not current_user.es_admin: abort(403)
    db.session.delete(c); db.session.commit()
    return jsonify({"ok":True})

# ═════════════════════════════════════════════════════════════════════════════
#  NOTIFICACIONES
# ═════════════════════════════════════════════════════════════════════════════
@app.route("/notificaciones")
@login_required
def notificaciones():
    notifs = (Notificacion.query.filter_by(usuario_id=current_user.id)
              .order_by(Notificacion.fecha.desc()).limit(50).all())
    return render_template("notificaciones.html", notificaciones=notifs)

@app.route("/api/notificaciones")
@login_required
def api_notificaciones():
    notifs = (Notificacion.query.filter_by(usuario_id=current_user.id, leida=False)
              .order_by(Notificacion.fecha.desc()).limit(10).all())
    total_no_leidas = Notificacion.query.filter_by(
        usuario_id=current_user.id, leida=False).count()
    return jsonify({
        "total": total_no_leidas,
        "items": [{"id":n.id,"tipo":n.tipo,"titulo":n.titulo,
                   "texto":n.texto,"url":n.url,
                   "fecha":n.fecha.strftime("%H:%M")} for n in notifs]
    })

@app.route("/api/notificaciones/leer", methods=["POST"])
@login_required
def notif_leer_todas():
    Notificacion.query.filter_by(usuario_id=current_user.id, leida=False)\
        .update({"leida":True})
    db.session.commit()
    return jsonify({"ok":True})

@app.route("/api/notificaciones/<int:nid>/leer", methods=["POST"])
@login_required
def notif_leer_una(nid):
    n = Notificacion.query.get_or_404(nid)
    if n.usuario_id == current_user.id:
        n.leida = True; db.session.commit()
    return jsonify({"ok":True})

# ═════════════════════════════════════════════════════════════════════════════
#  CHAT
# ═════════════════════════════════════════════════════════════════════════════
@app.route("/chat")
@login_required
def chat():
    """Lista de conversaciones del usuario."""
    from sqlalchemy import or_, func

    # Determinar con quién puede chatear según el rol
    if current_user.es_comercial:
        # Puede chatear con supervisores y admins
        interlocutores = Usuario.query.filter(
            Usuario.rol.in_(["supervisor","admin"]), Usuario.activo==True,
            Usuario.id != current_user.id).all()
    elif current_user.es_supervisor:
        # Puede chatear con comerciales y admins
        interlocutores = Usuario.query.filter(
            Usuario.rol.in_(["comercial","admin"]), Usuario.activo==True,
            Usuario.id != current_user.id).all()
    else:  # admin
        # Puede chatear con todos
        interlocutores = Usuario.query.filter(
            Usuario.id != current_user.id, Usuario.activo==True).all()

    # Último mensaje con cada interlocutor
    convs = []
    for u in interlocutores:
        ult = (MensajeChat.query.filter(
            or_(
                (MensajeChat.de_id==current_user.id) & (MensajeChat.para_id==u.id),
                (MensajeChat.de_id==u.id) & (MensajeChat.para_id==current_user.id)
            )).order_by(MensajeChat.fecha.desc()).first())
        no_leidos = MensajeChat.query.filter_by(
            de_id=u.id, para_id=current_user.id, leido=False).count()
        convs.append({"usuario":u, "ultimo":ult, "no_leidos":no_leidos})

    # Ordenar por último mensaje (más reciente arriba)
    convs.sort(key=lambda x: x["ultimo"].fecha if x["ultimo"] else datetime.min, reverse=True)

    uid_activo = request.args.get("con", type=int)
    mensajes = []
    interlocutor_activo = None
    if uid_activo:
        interlocutor_activo = db.session.get(Usuario, uid_activo)
        from sqlalchemy import or_
        mensajes = (MensajeChat.query.filter(
            or_(
                (MensajeChat.de_id==current_user.id) & (MensajeChat.para_id==uid_activo),
                (MensajeChat.de_id==uid_activo) & (MensajeChat.para_id==current_user.id)
            )).order_by(MensajeChat.fecha.asc()).all())
        # Marcar como leídos
        MensajeChat.query.filter_by(de_id=uid_activo, para_id=current_user.id, leido=False)\
            .update({"leido":True})
        db.session.commit()

    return render_template("chat.html", convs=convs, mensajes=mensajes,
                           interlocutor=interlocutor_activo)

@app.route("/chat/enviar", methods=["POST"])
@login_required
def chat_enviar():
    data = request.json or {}
    para_id = data.get("para_id")
    texto   = (data.get("texto") or "").strip()
    if not para_id or not texto:
        return jsonify({"error":"faltan datos"}), 400

    dest = db.session.get(Usuario, para_id)
    if not dest: return jsonify({"error":"usuario no encontrado"}), 404

    msg = MensajeChat(de_id=current_user.id, para_id=para_id, texto=texto)
    db.session.add(msg)

    # Notificar al destinatario
    crear_notif(para_id, "chat",
        f"💬 Mensaje de {current_user.nombre}",
        texto[:80],
        url_for("chat", con=current_user.id))

    db.session.commit()
    return jsonify({
        "id": msg.id, "texto": msg.texto,
        "de": current_user.nombre,
        "fecha": msg.fecha.strftime("%H:%M"),
        "mio": True
    })

@app.route("/chat/mensajes/<int:uid>")
@login_required
def chat_mensajes(uid):
    """Polling de mensajes nuevos desde un timestamp."""
    desde_id = request.args.get("desde_id", 0, type=int)
    from sqlalchemy import or_
    msgs = (MensajeChat.query.filter(
        MensajeChat.id > desde_id,
        or_(
            (MensajeChat.de_id==current_user.id) & (MensajeChat.para_id==uid),
            (MensajeChat.de_id==uid) & (MensajeChat.para_id==current_user.id)
        )).order_by(MensajeChat.fecha.asc()).all())
    # Marcar como leídos
    for m in msgs:
        if m.para_id == current_user.id and not m.leido:
            m.leido = True
    db.session.commit()
    return jsonify([{
        "id":m.id,"texto":m.texto,"de":m.de.nombre,
        "fecha":m.fecha.strftime("%H:%M"),"mio":(m.de_id==current_user.id)
    } for m in msgs])

@app.route("/api/chat/no_leidos")
@login_required
def chat_no_leidos():
    n = MensajeChat.query.filter_by(para_id=current_user.id, leido=False).count()
    return jsonify({"total": n})

# ═════════════════════════════════════════════════════════════════════════════
#  WORKERS
# ═════════════════════════════════════════════════════════════════════════════
def _run_scrape(flask_app, asig_id):
    with flask_app.app_context():
        asig = db.session.get(Asignacion, asig_id)
        if not asig: return

        def upd(estado, pct, msg):
            asig.estado=estado; asig.progreso=pct; asig.mensaje=msg
            db.session.commit()

        # Esperar slot global para no disparar bloqueos 429 por concurrencia.
        upd("pendiente", 1, "En cola anti-bloqueo… esperando turno de scraping")
        acquired = False
        t0 = time.time()
        while not acquired:
            acquired = _SCRAPE_SLOT.acquire(timeout=2.0)
            if not acquired:
                waited = int(time.time() - t0)
                asig.estado = "pendiente"
                asig.progreso = 1
                asig.mensaje = f"En cola anti-bloqueo… {waited}s"
                db.session.commit()

        upd("scrapeando", 5, "Iniciando scraping…")

        def on_progress(pct, msg):
            asig.progreso = 5 + int(pct * 0.75)
            asig.mensaje  = msg
            db.session.commit()

        try:
            max_429 = max(0, int(getattr(config, "SCRAPE_RETRY_429_ATTEMPTS", 4)))
            max_soft = max(0, int(getattr(config, "SCRAPE_RESCUE_ATTEMPTS", 2)))
            intento_429 = 0
            intento_soft = 0
            leads_raw = pool = None
            err = None
            meta = {}
            nacional_exhaustivo = False
            max_low_volume = max_soft

            while True:
                paginas_cfg = None if (asig.paginas is None or int(asig.paginas or 0) <= 0) else asig.paginas
                nacional_exhaustivo = (paginas_cfg is None) and not (asig.provincia or "").strip()
                max_low_volume = (2 if nacional_exhaustivo else max_soft)
                max_429_local = min(max_429, 2) if nacional_exhaustivo else max_429
                prefer_full_portal = nacional_exhaustivo
                delay_eff = min(float(config.SCRAPE_DELAY_SECONDS), 1.2) if nacional_exhaustivo else float(config.SCRAPE_DELAY_SECONDS)
                leads_raw, pool, err, meta = scrape_cnae(
                    cnae=asig.cnae, provincia=asig.provincia or None,
                    paginas=paginas_cfg, delay=delay_eff,
                    on_progress=on_progress,
                    prefer_full_portal=prefer_full_portal,
                )
                if err and "429" in str(err) and intento_429 < max_429_local:
                    intento_429 += 1
                    espera = (min(45, 10 + 12 * intento_429) if nacional_exhaustivo
                              else min(180, 35 + 25 * intento_429))
                    upd("pendiente", 2, (
                        f"Bloqueo temporal del portal (429). "
                        f"Reintento automático {intento_429}/{max_429_local} en {espera}s…"
                    ))
                    time.sleep(espera)
                    upd("scrapeando", 5, "Reanudando scraping tras espera anti-bloqueo…")
                    continue
                if err and "429" in str(err) and nacional_exhaustivo:
                    # Tras agotar reintentos al portal principal, activar rescate fallback.
                    leads_raw, pool, err, meta = scrape_cnae(
                        cnae=asig.cnae, provincia=asig.provincia or None,
                        paginas=paginas_cfg, delay=config.SCRAPE_DELAY_SECONDS,
                        on_progress=on_progress,
                        prefer_full_portal=False,
                    )

                soft_patterns = (
                    "No fue posible obtener resultados del portal principal ni del fallback web",
                    "No se pudieron obtener resultados en el rescate nacional por provincias",
                    "No se pudo cargar la fuente fallback empresascif",
                    "Tiempo límite alcanzado",
                )
                if err and intento_soft < max_soft and any(p in str(err) for p in soft_patterns):
                    intento_soft += 1
                    espera = min(120, 20 + 18 * intento_soft)
                    upd("pendiente", 2, (
                        f"Fuentes web temporalmente inestables. "
                        f"Reintento inteligente {intento_soft}/{max_soft} en {espera}s…"
                    ))
                    time.sleep(espera)
                    upd("scrapeando", 5, "Reintentando scraping con rutas de rescate…")
                    continue
                fuente = (meta or {}).get("fuente")
                low_volume_fallback = (
                    not err
                    and nacional_exhaustivo
                    and fuente in (
                        "fallback_ddg",
                        "fallback_search_empresascif",
                        "fallback_empresascif",
                        "fallback_search_empresascif_nacional",
                        "fallback_empresascif_nacional",
                    )
                    and len(leads_raw or []) < 80
                )
                if low_volume_fallback and intento_soft < max_low_volume:
                    intento_soft += 1
                    espera = (min(35, 8 + 6 * intento_soft) if nacional_exhaustivo
                              else min(120, 24 + 20 * intento_soft))
                    upd("pendiente", 2, (
                        f"Volumen bajo ({len(leads_raw or [])} leads). "
                        f"Reintento inteligente {intento_soft}/{max_low_volume} en {espera}s para ampliar cobertura…"
                    ))
                    time.sleep(espera)
                    upd("scrapeando", 5, "Reintentando scraping nacional para ampliar leads…")
                    continue
                break
        except Exception as e:
            upd("error", 0, f"Excepción: {e}")
            return
        finally:
            if acquired:
                _SCRAPE_SLOT.release()

        if err:
            err_txt = str(err or "")
            tolerables = (
                "No fue posible obtener resultados del portal principal ni del fallback web",
                "No se pudieron obtener resultados en el rescate nacional por provincias",
                "Tiempo límite alcanzado",
            )
            if any(t in err_txt for t in tolerables):
                asig.total_leads = 0
                asig.estado = "completada"
                asig.progreso = 100
                asig.mensaje = (
                    "⚠️ Sin resultados en este intento por bloqueo temporal de fuentes externas. "
                    "Puedes reintentar en unos minutos."
                )
                asig.fecha_completada = utcnow()
                db.session.commit()
                return
            upd("error", 0, err); return
        if not leads_raw:
            fuente = (meta or {}).get("fuente")
            actividad_count = (meta or {}).get("actividad_count")
            fichas_validadas = (meta or {}).get("fichas_validadas")

            asig.total_leads = 0
            asig.estado = "completada"
            asig.progreso = 100
            if fuente == "fallback_empresascif":
                extra = ""
                if actividad_count:
                    extra += f" · Ref. portal: ~{actividad_count} empresas"
                if fichas_validadas:
                    extra += f" · Fichas validadas: {fichas_validadas}"
                asig.mensaje = (
                    f"⚠️ 0 leads para CNAE {asig.cnae} en {asig.provincia or 'España'} "
                    f"(fallback anti-bloqueo activo){extra}"
                )
            else:
                asig.mensaje = f"⚠️ 0 leads para CNAE {asig.cnae} en {asig.provincia or 'España'}"
            asig.fecha_completada = utcnow()
            db.session.commit()
            return

        upd("scrapeando", 82, f"Guardando {len(leads_raw)} leads…")

        # Importar función de domain guessing para pre-enriquecimiento rápido
        from enrichment import enrich_from_domain_guess

        # Crear leads + competidores en BD
        # + Pre-enriquecimiento instantáneo vía Clearbit (sin delay, muy rápido)
        for e in leads_raw:
            gerente_scrape = _norm_gerente(e.get("gerente")) if e.get("gerente") else None
            lead = Lead(
                asignacion_id=asig.id, comercial_id=asig.comercial_id,
                nombre=e["nombre"], cnae=e["cnae"], provincia=e["provincia"],
                posicion_nacional=e["posicion"], evolucion=e["evolucion"],
                tendencia=e["tendencia"], facturacion_num=e["facturacion_num"],
                facturacion_raw=e["facturacion_raw"], url_ficha=e["url"],
                gerente=gerente_scrape,
                estado="nuevo",
            )
            # Intento rápido de domain guessing durante el guardado
            try:
                quick = enrich_from_domain_guess(e["nombre"])
                if quick.get("web"): lead.web = _norm_web(quick["web"])
                if quick.get("email"): lead.email = _norm_email(quick["email"])
                if quick.get("telefono"): lead.telefono = _norm_phone(quick["telefono"])
                if quick.get("direccion"): lead.direccion = (quick["direccion"] or "").strip()[:300] or None
            except Exception:
                pass
            db.session.add(lead)
            db.session.flush()  # obtener lead.id

            # Calcular 3 competidores
            comps = calcular_competidores(e, pool, ratio_min=3, ratio_max=30, n=3)
            for i, c in enumerate(comps):
                ratio = None
                if e.get("facturacion_num") and c.get("facturacion_num"):
                    ratio = round(c["facturacion_num"] / e["facturacion_num"], 1)
                db.session.add(Competidor(
                    lead_id=lead.id, orden=i+1,
                    nombre=c["nombre"], cnae=c.get("cnae", ""),
                    provincia=c["provincia"],
                    facturacion_raw=c["facturacion_raw"],
                    facturacion_num=c["facturacion_num"],
                    posicion=c["posicion"], tendencia=c["tendencia"],
                    ratio=ratio,
                    misma_provincia=(c["provincia"]==e["provincia"]),
                    url_ficha=c.get("url",""),
                ))

        asig.total_leads = len(leads_raw)
        asig.estado = "completada"
        asig.progreso = 100

        # Mensaje informativo con contexto de rastreo
        pags_reales = meta.get("paginas_reales", 0)
        agotado     = meta.get("agotado", False)
        fuente = meta.get("fuente")
        fichas_validadas = meta.get("fichas_validadas")
        if fuente in ("fallback_ddg", "fallback_empresascif", "fallback_search_empresascif", "fallback_search_empresascif_nacional", "fallback_empresascif_nacional"):
            base = (
                f"✅ {len(leads_raw)} leads · Fallback anti-bloqueo web activo "
                f"(portal principal bloqueado temporalmente)"
            )
            if fichas_validadas:
                base += f" · {fichas_validadas} fichas validadas"
            asig.mensaje = base
        elif agotado:
            asig.mensaje = (
                f"✅ {len(leads_raw)} leads · "
                f"{pags_reales} páginas rastreadas · "
                f"CNAE agotado"
            )
        else:
            asig.mensaje = (
                f"✅ {len(leads_raw)} leads · "
                f"{pags_reales} páginas rastreadas"
            )
        asig.fecha_completada = utcnow()

        # Notificar al comercial y admin (dentro del mismo app_context)
        crear_notif(asig.comercial_id, "success",
            f"✅ Asignación lista: {len(leads_raw)} leads",
            f"CNAE {asig.cnae} · {asig.provincia or 'Nacional'}",
            "/kanban")
        if asig.creado_por_id != asig.comercial_id:
            crear_notif(asig.creado_por_id, "info",
                f"Scraping completado: {asig.cnae}/{asig.provincia or 'España'}",
                f"{len(leads_raw)} leads para {asig.comercial.nombre}",
                "/admin/asignaciones")

        # ⚡ Commit COMPLETADA antes de lanzar enrichment
        # Así el progreso aparece al 100% inmediatamente
        db.session.commit()

        # Lanzar enriquecimiento en background (no bloquea el progreso)
        lead_ids = [l.id for l in asig.leads.all()]
        threading.Thread(target=_enrich_batch, args=(flask_app, lead_ids),
                         daemon=True).start()


def _enrich_one(flask_app, lead_id):
    with flask_app.app_context():
        lead = db.session.get(Lead, lead_id)
        if not lead: return
        try:
            datos = enrich_lead({"nombre":lead.nombre, "provincia":lead.provincia,
                                  "url":lead.url_ficha})

            # Normalización de calidad de datos antes de guardar.
            if datos.get("telefono"):
                datos["telefono"] = _norm_phone(datos.get("telefono"))
            if datos.get("email"):
                datos["email"] = _norm_email(datos.get("email"))
            if datos.get("web"):
                datos["web"] = _norm_web(datos.get("web"))
            if datos.get("direccion"):
                datos["direccion"] = (datos.get("direccion") or "").strip()[:300]
            if datos.get("gerente"):
                datos["gerente"] = _norm_gerente(datos.get("gerente"))

            for k in ("telefono", "email", "web", "direccion", "gerente", "licita"):
                v = datos.get(k)
                if v:
                    setattr(lead, k, v)

            comp = _lead_completeness_map({
                "telefono": lead.telefono or "",
                "email": lead.email or "",
                "web": lead.web or "",
                "direccion": lead.direccion or "",
                "gerente": lead.gerente or "",
            })
            lead.enriquecido = comp >= max(1, int(getattr(config, "ENRICHMENT_MIN_COMPLETENESS", 2)))
            db.session.commit()
        except Exception as e:
            print(f"[enrich_one] {lead_id}: {e}")


def _enrich_batch(flask_app, lead_ids):
    """
    Enriquecimiento paralelo máximo.
    - 10 robots simultáneos = ~10x más rápido que la versión serie.
    - Inicio escalonado: cada robot arranca 0.5s después del anterior
      para evitar 10 peticiones simultáneas a Bing que dispararían el bloqueo.
    - Jitter aleatorio por robot entre búsquedas para imitar humanos.
    - 200 leads @ 10 workers ≈ ~3-4 minutos (vs. ~35 min en serie).
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import time, random

    WORKERS = max(1, int(getattr(config, "ENRICHMENT_MAX_WORKERS", 10)))
    START_STAGGER = 0.5  # segundos entre arranque de cada worker

    def _worker(args):
        lid, worker_idx = args
        # Arranque escalonado: worker 0 empieza ya, worker 9 empieza en 4.5s
        time.sleep(worker_idx * START_STAGGER)
        try:
            _enrich_one(flask_app, lid)
        except Exception as e:
            print(f"[enrich_batch] {lid}: {e}")
        # Jitter aleatorio al terminar cada lead para no sincronizarse
        time.sleep(random.uniform(0.5, 1.5))

    args_list = [(lid, i % WORKERS) for i, lid in enumerate(lead_ids)]
    with ThreadPoolExecutor(max_workers=WORKERS) as executor:
        futures = {executor.submit(_worker, a): a for a in args_list}
        for f in as_completed(futures):
            pass  # errores manejados dentro de _worker

    # Segunda pasada opcional sobre leads incompletos para mejorar cobertura.
    if not bool(getattr(config, "ENRICHMENT_SECOND_PASS", True)):
        return

    with flask_app.app_context():
        min_comp = max(1, int(getattr(config, "ENRICHMENT_MIN_COMPLETENESS", 2)))
        pass2_ids = []
        for lid in lead_ids:
            lead = db.session.get(Lead, lid)
            if not lead:
                continue
            comp = _lead_completeness_map({
                "telefono": lead.telefono or "",
                "email": lead.email or "",
                "web": lead.web or "",
                "direccion": lead.direccion or "",
                "gerente": lead.gerente or "",
            })
            if comp < min_comp:
                pass2_ids.append(lid)

    if not pass2_ids:
        return

    WORKERS_2 = max(1, min(6, WORKERS // 2 or 1))
    args_list_2 = [(lid, i % WORKERS_2) for i, lid in enumerate(pass2_ids)]
    with ThreadPoolExecutor(max_workers=WORKERS_2) as executor:
        futures = {executor.submit(_worker, a): a for a in args_list_2}
        for f in as_completed(futures):
            pass


# ═════════════════════════════════════════════════════════════════════════════
#  EXCEL
# ═════════════════════════════════════════════════════════════════════════════
def _build_excel(leads):
    THIN = Border(*[Side(style="thin", color="D0D5DD")]*4)
    wb = Workbook(); ws = wb.active; ws.title = "Leads"
    ws.sheet_properties.tabColor = "DC2626"
    
    # 1. Título principal
    ws.merge_cells("A1:Z1")
    c = ws.cell(1,1,f"RADAR CRM · Leads de Alta Calidad · Generado el {datetime.now():%d/%m/%Y}")
    c.font = Font(name="Inter", bold=True, size=15, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="111827")  # Very dark grey
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 2. Definir los Super-Grupos (Fila 3)
    grupos = [
        ("A3:F3", "🏢 EMPRESA Y RANKING", "1E3A8A"),      # Azul oscuro
        ("G3:L3", "📞 DATOS DE CONTACTO", "065F46"),      # Verde esmeralda oscuro
        ("M3:N3", "🎯 CRM", "991B1B"),                    # Rojo oscuro
        ("O3:R3", "🏆 COMPETIDOR #1", "5B21B6"),          # Púrpura oscuro
        ("S3:V3", "🏆 COMPETIDOR #2", "4C1D95"),          # Púrpura más oscuro
        ("W3:Z3", "🏆 COMPETIDOR #3", "3B0764"),          # Púrpura ultra oscuro
    ]
    ws.row_dimensions[3].height = 22
    for rango, texto, color in grupos:
        ws.merge_cells(rango)
        start_col = ws[rango.split(":")[0]]
        start_col.value = texto
        start_col.font = Font(name="Inter", bold=True, size=11, color="FFFFFF")
        start_col.fill = PatternFill("solid", start_color=color)
        start_col.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws[rango]:
            for cell in row:
                cell.border = Border(top=Side(style="thin", color=color), bottom=Side(style="thin", color=color))

    # 3. Cabeceras de Columna (Fila 4)
    hdrs = ["Empresa","CNAE","Provincia","Posición","Facturación","Tendencia",
            "Teléfono","Email","Web","Dirección","Gerente","Licita",
            "Estado","Comercial",
            "Nombre","Provincia","Facturación","Pos.",
            "Nombre","Provincia","Facturación","Pos.",
            "Nombre","Provincia","Facturación","Pos."]
            
    ws.row_dimensions[4].height = 20
    for i,h in enumerate(hdrs,1):
        cell = ws.cell(4,i,h)
        cell.font = Font(name="Inter", bold=True, color="1F2937") # Gris oscuro
        cell.fill = PatternFill("solid", start_color="F3F4F6") # Gris muy claro
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN

    # 4. Inserción de Datos (Desde Fila 5)
    for r,l in enumerate(leads, 5):
        bg = "F9FAFB" if r%2==0 else "FFFFFF" # Cebra muy sutil
        comps = list(l.competidores.all()) + [None, None, None]
        
        def _get_comp_data(c):
            return [c.nombre, c.provincia, c.facturacion_raw, c.posicion] if c else ["","","",""]
            
        c1 = _get_comp_data(comps[0])
        c2 = _get_comp_data(comps[1])
        c3 = _get_comp_data(comps[2])
        
        vals = [l.nombre,l.cnae,l.provincia,l.posicion_nacional,l.facturacion_raw,
                l.tendencia,l.telefono,l.email,l.web,l.direccion,l.gerente,l.licita,
                l.estado,l.comercial.nombre if l.comercial else ""] + c1 + c2 + c3
                
        for i,v in enumerate(vals,1):
            val = v or ""
            if isinstance(val, str):
                import re
                val = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', val)
            cell = ws.cell(r,i,val)
            cell.font = Font(name="Inter", size=10, color="374151")
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(vertical="center", horizontal="left" if i in (1,8,9,10,11,15,19,23) else "center")
            cell.border = THIN

    # 5. Ajustes Finales de UI Excel
    ws.auto_filter.ref = f"A4:Z{ws.max_row}" # Habilitar filtros
    ws.freeze_panes = "C5" # Congelar las 2 primeras columnas (Empresa y CNAE) y las filas de cabecera
    
    widths = [45, 7, 16, 10, 20, 12, 14, 32, 32, 45, 26, 9, 14, 18,
              35, 16, 20, 8, 35, 16, 20, 8, 35, 16, 20, 8]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w
        
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ═════════════════════════════════════════════════════════════════════════════
#  CONTEXT / FILTERS
# ═════════════════════════════════════════════════════════════════════════════
@app.context_processor
def inject_globals():
    notif_count = chat_count = 0
    if current_user.is_authenticated:
        try:
            notif_count = Notificacion.query.filter_by(
                usuario_id=current_user.id, leida=False).count()
            chat_count = MensajeChat.query.filter_by(
                para_id=current_user.id, leido=False).count()
        except Exception:
            pass  # Tablas aún no creadas o error de BD
    return {"KANBAN_ESTADOS":config.KANBAN_ESTADOS, "ROLES":config.ROLES,
            "notif_count":notif_count, "chat_count":chat_count}

@app.template_filter("fmt_fecha")
def fmt_fecha(f): return f.strftime("%d/%m/%Y %H:%M") if f else ""

@app.template_filter("fmt_num")
def fmt_num(n):
    if n is None: return "—"
    try: return f"{int(n):,}".replace(",",".")
    except: return str(n)

@app.errorhandler(403)
def e403(e): return render_template("403.html"), 403

if __name__ == "__main__":
    print("\n"+"═"*52)
    print("  RADAR CRM v2  ·  http://localhost:5000")
    print("═"*52+"\n")
    app.run(debug=False, port=5000, host="0.0.0.0")
